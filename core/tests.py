import json
import shutil
import tempfile
from datetime import timedelta
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote
from unittest.mock import Mock, patch

from django.conf import settings
from django.contrib.auth.hashers import make_password
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from django.utils import timezone
import jwt
from openpyxl import load_workbook
from rest_framework.test import APIClient, APITestCase

from algorithm.adapter import append_log
from core.models import (
    ApsArchive, ApsArchiveItem, SolveTask, SysUser, TaskImportRecord, UploadFile, UploadFileItem,
)
from services.excel_service import parse_aps_file
from services.solve_match_service import (
    REASON_MISSING_BOTH,
    REASON_MISSING_PRODUCT,
    REASON_MISSING_SPEC,
    compare_task_plan_with_aps,
)


class ApiIntegrationTests(APITestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.temp_dir = tempfile.mkdtemp(prefix="tdsms-tests-")
        cls.override = override_settings(
            TASK_UPLOAD_ROOT=str(Path(cls.temp_dir) / "uploads"),
            RESULT_EXCEL_ROOT=str(Path(cls.temp_dir) / "results"),
            ALGORITHM_LOG_ROOT=str(Path(cls.temp_dir) / "logs"),
            ALGORITHM_RUN_ROOT=str(Path(cls.temp_dir) / "algorithm_runs"),
            ALGORITHM_INPUT_ROOT=str(Path(cls.temp_dir) / "algorithm_inputs"),
            USE_MOCK_ALGORITHM=False,
        )
        cls.override.enable()

    @classmethod
    def tearDownClass(cls):
        cls.override.disable()
        shutil.rmtree(cls.temp_dir, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        self.user = SysUser.objects.create(username="user01", password=make_password("123456"), realName="张三")
        self.other = SysUser.objects.create(username="user02", password=make_password("123456"))
        self.admin = SysUser.objects.create(username="admin", password=make_password("admin123"), role="admin")
        self.client = APIClient()

    def login(self, username="user01", password="123456"):
        response = self.client.post("/tdsms/auth/login", {"username": username, "password": password}, format="json")
        self.assertEqual(response.status_code, 200, response.data)
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {response.data['data']['token']}")
        return response

    def create_archive(self):
        self.login()
        template = Path(settings.MEDIA_ROOT) / "templates" / "APS排产信息模板.xlsx"
        with template.open("rb") as stream:
            response = self.client.post("/tdsms/aps/create", {"archiveName": "方案一", "file": stream}, format="multipart")
        self.assertEqual(response.status_code, 200, response.data)
        return ApsArchive.objects.get(archiveId=response.data["data"]["archiveId"])

    def create_task(self):
        archive = self.create_archive()
        template = Path(settings.MEDIA_ROOT) / "templates" / "药业车间分解编排计划表模板.xlsx"
        with template.open("rb") as stream:
            response = self.client.post("/tdsms/task/import", {"apsArchiveId": archive.archiveId, "remark": "测试", "file": stream}, format="multipart")
        self.assertEqual(response.status_code, 200, response.data)
        return TaskImportRecord.objects.get(taskId=response.data["data"]["taskId"])

    def solvable_department(self, task):
        departments = task.file.items.filter(isDeleted=0).values_list(
            "departmentName", flat=True,
        ).distinct()
        for department in departments:
            if compare_task_plan_with_aps(task, department)["matchedCount"]:
                return department
        self.fail("测试模板中不存在可与APS档案匹配的部门")

    def create_solve_record(self, **solve_kwargs):
        self.login()
        archive = ApsArchive.objects.create(archiveName="测试档案", createdBy=self.user)
        upload = UploadFile.objects.create(
            originalName="plan.xlsx",
            fileName="plan.xlsx",
            filePath="plan.xlsx",
            uploadUser=self.user,
            parseStatus=1,
        )
        task = TaskImportRecord.objects.create(
            apsArchive=archive,
            file=upload,
            importStatus=1,
            createdBy=self.user,
        )
        fields = {
            "task": task,
            "createdUser": self.user,
            "inputParams": {},
            "solveStatus": 1,
        }
        fields.update(solve_kwargs)
        return SolveTask.objects.create(**fields)

    def test_authentication_and_admin_permission(self):
        response = self.login()
        self.assertEqual(response.data["data"]["userInfo"]["role"], "user")
        denied = self.client.get("/tdsms/admin/query")
        self.assertEqual(denied.status_code, 403)
        self.client.credentials()
        self.login("admin", "admin123")
        created = self.client.post("/tdsms/admin/create", {"username": "newuser", "password": "123456", "validDays": 30}, format="json")
        self.assertEqual(created.status_code, 200, created.data)

    def test_single_login_logout_and_old_token_invalidation(self):
        first = self.client.post(
            "/tdsms/auth/login",
            {"username": "user01", "password": "123456"},
            format="json",
        )
        self.assertEqual(first.status_code, 200, first.data)
        token = first.data["data"]["token"]
        self.user.refresh_from_db()
        self.assertEqual(self.user.loginToken, token)

        repeated = self.client.post(
            "/tdsms/auth/login",
            {"username": "user01", "password": "123456"},
            format="json",
        )
        self.assertEqual(repeated.status_code, 409, repeated.data)
        self.assertEqual(repeated.data["message"], "当前账号已经登录，请勿重复登录")

        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")
        logged_out = self.client.post("/tdsms/auth/logout", {}, format="json")
        self.assertEqual(logged_out.status_code, 200, logged_out.data)
        self.user.refresh_from_db()
        self.assertIsNone(self.user.loginToken)

        rejected = self.client.get("/tdsms/aps/listQuery")
        self.assertEqual(rejected.status_code, 403, rejected.data)

        self.client.credentials()
        logged_in_again = self.client.post(
            "/tdsms/auth/login",
            {"username": "user01", "password": "123456"},
            format="json",
        )
        self.assertEqual(logged_in_again.status_code, 200, logged_in_again.data)
        self.assertNotEqual(logged_in_again.data["data"]["token"], token)

    def test_expired_stored_token_is_replaced_on_login(self):
        expired_token = jwt.encode(
            {
                "userId": self.user.userId,
                "username": self.user.username,
                "exp": int((timezone.now() - timedelta(minutes=1)).timestamp()),
            },
            settings.JWT_SECRET_KEY,
            algorithm=settings.JWT_ALGORITHM,
        )
        self.user.loginToken = expired_token
        self.user.save(update_fields=["loginToken"])

        response = self.client.post(
            "/tdsms/auth/login",
            {"username": "user01", "password": "123456"},
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.user.refresh_from_db()
        self.assertNotEqual(self.user.loginToken, expired_token)

    def test_disabling_user_clears_login_token_immediately(self):
        user_client = APIClient()
        login = user_client.post(
            "/tdsms/auth/login",
            {"username": "user01", "password": "123456"},
            format="json",
        )
        self.assertEqual(login.status_code, 200, login.data)
        user_token = login.data["data"]["token"]

        self.login("admin", "admin123")
        disabled = self.client.post(
            "/tdsms/admin/statusUpdate",
            {"userId": self.user.userId, "status": 0},
            format="json",
        )
        self.assertEqual(disabled.status_code, 200, disabled.data)
        self.user.refresh_from_db()
        self.assertEqual(self.user.status, 0)
        self.assertIsNone(self.user.loginToken)

        user_client.credentials(HTTP_AUTHORIZATION=f"Bearer {user_token}")
        rejected = user_client.get("/tdsms/aps/listQuery")
        self.assertEqual(rejected.status_code, 403, rejected.data)

    def test_template_downloads(self):
        self.login()
        cases = [
            ("/tdsms/aps/template", "APS排产信息模板.xlsx"),
            ("/tdsms/task/template", "药业车间分解编排计划表模板.xlsx"),
        ]
        for url, filename in cases:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200)
            self.assertIn(filename, unquote(response.headers["Content-Disposition"]))
            self.assertGreater(sum(len(chunk) for chunk in response.streaming_content), 0)

    def test_aps_and_plan_import_history_reuse_and_isolation(self):
        task = self.create_task()
        self.assertGreater(task.file.items.count(), 700)
        self.assertTrue(Path(task.file.filePath).is_file())
        history = self.client.get("/tdsms/task/historyQuery", {"page": 1, "pageSize": 10})
        self.assertEqual(history.status_code, 200)
        self.assertEqual(history.data["data"]["total"], 1)
        reused = self.client.post("/tdsms/tasks/historyImport", {"taskId": task.taskId}, format="json")
        self.assertEqual(reused.status_code, 200, reused.data)
        self.assertEqual(reused.data["data"]["sourceType"], 2)
        self.client.credentials()
        self.login("user02", "123456")
        hidden = self.client.get("/tdsms/task/detailQuery", {"importId": task.taskId})
        self.assertEqual(hidden.status_code, 404)

    def test_plan_detail_json_filters_and_single_filter_option_query(self):
        task = self.create_solve_record().task
        first = UploadFileItem.objects.create(
            file=task.file,
            departmentName="302车间",
            materialCode="MAT-001",
            inventoryName="阿司匹林片",
            specification="100mg*24片",
            monthlyProductionPlan=100,
        )
        UploadFileItem.objects.create(
            file=task.file,
            departmentName="303车间",
            materialCode="MAT-002",
            inventoryName="维生素C片",
            specification="100mg*100片",
            monthlyProductionPlan=200,
        )

        payload = {
            "taskId": task.taskId,
            "keyword": "MAT",
            "departmentNames": ["302车间", "303车间"],
            "monthlyProductionPlans": [100, 200],
            "inventoryNames": ["阿司匹林片"],
            "page": 1,
            "pageSize": 10,
        }
        response = self.client.generic(
            "GET",
            "/tdsms/task/detailQuery",
            data=json.dumps(payload),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["data"]["total"], 1)
        self.assertEqual(response.data["data"]["records"][0]["itemId"], first.itemId)

        # 查询串重复传多选条件时必须保留全部值，不能被 QueryDict.dict() 压成最后一项
        multi_filters = self.client.get(
            "/tdsms/task/detailQuery",
            {
                "taskId": task.taskId,
                "departmentNames": ["302车间", "303车间"],
                "monthlyProductionPlans": [100, 200],
                "inventoryNames": ["阿司匹林片", "维生素C片"],
                "page": 1,
                "pageSize": 10,
            },
        )
        self.assertEqual(multi_filters.status_code, 200, multi_filters.data)
        self.assertEqual(multi_filters.data["data"]["total"], 2)
        self.assertEqual(
            {row["departmentName"] for row in multi_filters.data["data"]["records"]},
            {"302车间", "303车间"},
        )
        self.assertEqual(
            {row["inventoryName"] for row in multi_filters.data["data"]["records"]},
            {"阿司匹林片", "维生素C片"},
        )
        self.assertEqual(
            {row["monthlyProductionPlan"] for row in multi_filters.data["data"]["records"]},
            {100.0, 200.0},
        )

        options = self.client.post(
            "/tdsms/task/detailFilterOptions",
            {
                "taskId": task.taskId,
                "option": "departmentNames",
                "departmentNames": [],
                "monthlyProductionPlans": [],
                "inventoryNames": [],
            },
            format="json",
        )
        self.assertEqual(options.status_code, 200, options.data)
        self.assertEqual(options.data["data"], ["302车间", "303车间"])

        linked = self.client.post(
            "/tdsms/task/detailFilterOptions",
            {
                "taskId": task.taskId,
                "option": "departmentNames",
                "departmentNames": [],
                "monthlyProductionPlans": [100],
                "inventoryNames": ["阿司匹林片"],
            },
            format="json",
        )
        self.assertEqual(linked.status_code, 200, linked.data)
        self.assertEqual(linked.data["data"], ["302车间"])

        plans = self.client.post(
            "/tdsms/task/detailFilterOptions",
            {
                "taskId": task.taskId,
                "option": "monthlyProductionPlans",
                "departmentNames": ["303车间"],
                "monthlyProductionPlans": [],
                "inventoryNames": [],
            },
            format="json",
        )
        self.assertEqual(plans.status_code, 200, plans.data)
        self.assertEqual(plans.data["data"], [200.0])

        invalid = self.client.post(
            "/tdsms/task/detailFilterOptions",
            {
                "taskId": task.taskId,
                "option": "unknown",
                "departmentNames": [],
                "monthlyProductionPlans": [],
                "inventoryNames": [],
            },
            format="json",
        )
        self.assertEqual(invalid.status_code, 400)

    def test_aps_item_create_batch_delete_and_export(self):
        archive = self.create_archive()
        create_payload = {
            "archiveId": archive.archiveId,
            "productName": "接口测试品种",
            "packageSpecification": "10mg×10片",
            "mixingLine": "配料线A",
            "mixingBatchQuantity": 50,
            "mixingWorkerCount": 3,
            "centralizedProcurement": 1,
            "annualSales": 1200,
        }
        created = self.client.post("/tdsms/aps/itemCreate", create_payload, format="json")
        self.assertEqual(created.status_code, 200, created.data)
        self.assertEqual(created.data["message"], "APS明细新增成功")
        self.assertEqual(created.data["data"], {})
        created_item = archive.items.filter(productName="接口测试品种", packageSpecification="10mg×10片", isDeleted=0).get()
        self.assertEqual(created_item.centralizedProcurement, 1)

        second = dict(create_payload)
        second["packageSpecification"] = "20mg×10片"
        created_second = self.client.post("/tdsms/aps/itemCreate", second, format="json")
        self.assertEqual(created_second.status_code, 200, created_second.data)
        self.assertEqual(created_second.data["message"], "APS明细新增成功")

        failed_create = self.client.post("/tdsms/aps/itemCreate", {"archiveId": archive.archiveId}, format="json")
        self.assertEqual(failed_create.status_code, 400)
        self.assertEqual(failed_create.data["message"], "APS明细新增失败")
        self.assertEqual(failed_create.data["data"], {})

        update_payload = dict(create_payload)
        update_payload["itemId"] = created_item.itemId
        update_payload["productName"] = "接口测试品种-已改"
        updated = self.client.post("/tdsms/aps/itemUpdate", update_payload, format="json")
        self.assertEqual(updated.status_code, 200, updated.data)
        self.assertEqual(updated.data["message"], "APS明细修改成功")
        self.assertEqual(updated.data["data"], {})
        self.assertEqual(
            archive.items.get(itemId=created_item.itemId).productName,
            "接口测试品种-已改",
        )

        failed = self.client.post("/tdsms/aps/itemUpdate", create_payload, format="json")
        self.assertEqual(failed.status_code, 400)
        self.assertEqual(failed.data["message"], "APS明细修改失败")
        self.assertEqual(failed.data["data"], {})

        single_deleted = self.client.post("/tdsms/aps/itemDelete", {
            "batchMode": False,
            "archiveId": archive.archiveId,
            "itemId": created_item.itemId,
        }, format="json")
        self.assertEqual(single_deleted.status_code, 200, single_deleted.data)
        self.assertEqual(single_deleted.data["data"]["itemId"], created_item.itemId)

        deleted = self.client.post("/tdsms/aps/itemDelete", {
            "batchMode": True,
            "archiveId": archive.archiveId,
            "productNames": ["接口测试品种", "接口测试品种", "不存在品种"],
        }, format="json")
        self.assertEqual(deleted.status_code, 200, deleted.data)
        self.assertEqual(deleted.data["data"]["deletedProductCount"], 1)
        self.assertEqual(deleted.data["data"]["deletedItemCount"], 1)
        self.assertEqual(deleted.data["data"]["productNames"], ["接口测试品种"])

        exported = self.client.get("/tdsms/aps/export", {"archiveId": archive.archiveId})
        self.assertEqual(exported.status_code, 200)
        self.assertEqual(
            exported.headers["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        content = b"".join(exported.streaming_content)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        self.assertEqual(workbook.worksheets[0].max_column, 22)
        expected_count = archive.items.filter(isDeleted=0).count()
        self.assertEqual(workbook.worksheets[0].max_row, expected_count + 2)

        uploaded = SimpleUploadedFile("export.xlsx", content)
        parsed_rows = parse_aps_file(uploaded)
        self.assertEqual(len(parsed_rows), expected_count)
        self.assertNotIn("接口测试品种", {row["productName"] for row in parsed_rows})

        self.client.credentials()
        self.login("user02", "123456")
        hidden = self.client.get("/tdsms/aps/export", {"archiveId": archive.archiveId})
        self.assertEqual(hidden.status_code, 404)

    @patch("api.views.solve_views.stop_solver")
    @patch("api.views.solve_views.submit_solver")
    def test_solve_lifecycle_and_txt_logs(self, submit_solver_mock, stop_solver_mock):
        task = self.create_task()
        department = self.solvable_department(task)
        payload = {
            "taskId": task.taskId,
            "department": department,
            "scheduleMonth": "2025-06",
            "unmatchedItemPolicy": "SKIP",
            "productionRules": {
                "continuousRunLimitDays": 5.5,
                "cleaningDuration": {"majorCleaningDays": 0.5, "minorCleaningDays": 0.25, "periodicCleaningDays": 0.5},
                "shiftConversion": {"naturalDays": 1, "shiftCount": 2},
            },
            "personnelCapacity": {
                "mixing": 3,
                "tableting": 2,
                "coating": 2,
                "packaging": 4,
            },
            "solverTimeLimitMinutes": 20,
        }
        invalid_department = self.client.post(
            "/tdsms/solve/start",
            {**payload, "department": "不存在车间"},
            format="json",
        )
        self.assertEqual(invalid_department.status_code, 400, invalid_department.data)
        started = self.client.post("/tdsms/solve/start", payload, format="json")
        self.assertEqual(started.status_code, 200, started.data)
        solve_id = started.data["data"]["solveTaskId"]
        submit_solver_mock.assert_called_once_with(solve_id)
        self.assertEqual(
            SolveTask.objects.get(solveTaskId=solve_id).inputParams["personnelCapacity"],
            {"mixing": 3, "tableting": 2, "coating": 2, "packaging": 4},
        )
        self.assertEqual(
            SolveTask.objects.get(solveTaskId=solve_id).inputParams["department"],
            department,
        )
        append_log(solve_id, "正在计算搜索...")
        logs = self.client.get("/tdsms/solve/logs", {"solveTaskId": solve_id})
        self.assertIn(
            "正在计算搜索...",
            [record["logContent"] for record in logs.data["data"]],
        )
        for record in logs.data["data"]:
            if record["createTime"]:
                self.assertNotIn("T", record["createTime"])
        status_response = self.client.get("/tdsms/solve/query", {"solveTaskId": solve_id})
        self.assertEqual(status_response.data["data"]["solveStatus"], 0)
        stop_solver_mock.return_value = {"taskId": solve_id, "status": "STOPPING"}
        stopped = self.client.post("/tdsms/solve/stop", {"solveTaskId": solve_id}, format="json")
        self.assertEqual(stopped.status_code, 202)
        self.assertTrue(stopped.data["data"]["stopRequested"])
        self.assertEqual(SolveTask.objects.get(solveTaskId=solve_id).solveStatus, 0)

    def test_solve_logs_normalizes_create_time_and_returns_newest_first(self):
        solve = self.create_solve_record()
        run_dir = Path(settings.ALGORITHM_RUN_ROOT) / str(solve.solveTaskId)
        run_dir.mkdir(parents=True, exist_ok=True)
        (run_dir / "solver.log").write_text(
            "[2026-08-18T16:07:36] 任务创建: 36\n"
            "[2026-08-18T16:07:36] 求解子进程启动\n"
            "[2026-08-18T16:07:36] 开始计算排产方案...\n"
            "[2026-08-18T16:07:36] 已找到一个排产方案，正在进一步计算搜索...\n"
            "1. 读取输入数据...\n"
            "   初解[auto:priority] objective=0\n"
            "[2026-08-18T16:07:58] 任务已请求停止\n"
            "[2026-08-18T16:07:58] 收到停止请求，正在导出当前最佳方案\n",
            encoding="utf-8",
        )

        response = self.client.get("/tdsms/solve/logs", {"solveTaskId": solve.solveTaskId})

        self.assertEqual(response.status_code, 200, response.data)
        records = response.data["data"]
        self.assertEqual(
            [record["logContent"] for record in records],
            [
                "收到停止请求，正在导出当前最佳方案",
                "任务已请求停止",
                "已找到一个排产方案，正在进一步计算搜索...",
                "开始计算排产方案...",
                "求解子进程启动",
                "任务创建: 36",
            ],
        )
        self.assertEqual(
            [record["createTime"] for record in records],
            [
                "2026-08-18 16:07:58",
                "2026-08-18 16:07:58",
                "2026-08-18 16:07:36",
                "2026-08-18 16:07:36",
                "2026-08-18 16:07:36",
                "2026-08-18 16:07:36",
            ],
        )
        self.assertEqual([record["logId"] for record in records], [1, 2, 3, 4, 5, 6])

    def test_algorithm_adapter_prepares_inputs_and_maps_solver_parameters(self):
        from algorithm.adapter import submit_solver

        task = self.create_task()
        department = self.solvable_department(task)
        solve = SolveTask.objects.create(
            task=task,
            createdUser=self.user,
            inputParams={
                "department": department,
                "scheduleMonth": "2025-06",
                "unmatchedItemPolicy": "SKIP",
                "productionRules": {
                    "continuousRunLimitDays": 5.5,
                    "cleaningDuration": {
                        "majorCleaningDays": 0.5,
                        "minorCleaningDays": 0.25,
                        "periodicCleaningDays": 0.5,
                    },
                    "shiftConversion": {"naturalDays": 1, "shiftCount": 2},
                },
                "personnelCapacity": {
                    "mixing": 3,
                    "tableting": 2,
                    "coating": 2,
                    "packaging": 4,
                },
                "solverTimeLimitMinutes": 20,
            },
        )
        engine = Mock()
        engine.run_solver.return_value = {"taskId": str(solve.solveTaskId), "status": "RUNNING"}
        engine.query_solver_status.return_value = {
            "taskId": str(solve.solveTaskId),
            "status": "RUNNING",
            "startedAt": "2026-08-16T12:00:00",
            "finishedAt": None,
            "resultFiles": {},
        }
        with patch("algorithm.adapter._engine_service", return_value=engine), patch(
            "algorithm.adapter._start_monitor"
        ):
            result = submit_solver(solve.solveTaskId)

        self.assertEqual(result["taskId"], str(solve.solveTaskId))
        arguments = engine.run_solver.call_args.kwargs
        self.assertEqual(arguments["taskId"], solve.solveTaskId)
        self.assertEqual(arguments["planFile"], task.file.filePath)
        self.assertEqual(arguments["department"], department)
        self.assertEqual(arguments["mixing"], 3)
        self.assertTrue(Path(arguments["apsFile"]).is_file())
        workbook = load_workbook(arguments["apsFile"], read_only=True, data_only=True)
        self.assertGreater(workbook.worksheets[0].max_row, 2)
        workbook.close()
        solve.refresh_from_db()
        self.assertEqual(solve.solveStatus, 1)

        run_dir = Path(settings.ALGORITHM_RUN_ROOT) / str(solve.solveTaskId)
        run_dir.mkdir(parents=True, exist_ok=True)
        visual = run_dir / "可排产结果可视化.xlsx"
        visual.unlink(missing_ok=True)
        visual.write_bytes(b"result")
        intermediate = run_dir / "排产结果明细.xlsx"
        intermediate.write_bytes(b"temporary")
        engine.query_solver_status.return_value = {
            "taskId": str(solve.solveTaskId),
            "status": "SUCCESS",
            "startedAt": "2026-08-16T12:00:00",
            "finishedAt": "2026-08-16T12:20:00",
            "result": {"cp_status": "OPTIMAL"},
            "resultFiles": {
                "rawDetail": str(intermediate),
                "visualBoard": str(visual),
            },
        }
        with patch("algorithm.adapter._engine_service", return_value=engine):
            from algorithm.adapter import sync_solve_task

            sync_solve_task(solve)

        solve.refresh_from_db()
        self.assertEqual(solve.solveStatus, 2)
        self.assertEqual(solve.finishReason, 2)
        self.assertEqual(solve.resultFilePath, str(visual))
        self.assertTrue(visual.is_file())
        self.assertFalse(intermediate.exists())
        self.assertFalse(Path(arguments["apsFile"]).exists())

    def test_solve_match_check_applies_mappings_and_reports_unmatched_items(self):
        solve = self.create_solve_record()
        task = solve.task
        solve.solveStatus = 2
        solve.save(update_fields=["solveStatus"])
        task.file.items.update(isDeleted=1)
        task.apsArchive.items.update(isDeleted=1)
        plan_rows = [
            {
                "departmentName": "210车间",
                "materialCode": "M-001",
                "inventoryName": "阿司匹林肠溶片（过评）",
                "specification": "47.5mg×14粒×2板×400盒",
                "monthlyProductionPlan": 100,
            },
            {
                "departmentName": "210车间",
                "materialCode": "M-002",
                "inventoryName": "阿德福韦酯片",
                "specification": "10mg×28片×100瓶",
                "monthlyProductionPlan": 200,
            },
            {
                "departmentName": "210车间",
                "materialCode": "M-003",
                "inventoryName": "未维护药品",
                "specification": "20mg×10片",
                "monthlyProductionPlan": 300,
            },
            {
                "departmentName": "210车间",
                "materialCode": "M-003B",
                "inventoryName": "未维护药品",
                "specification": "20mg×10片",
                "monthlyProductionPlan": 10,
            },
            {
                "departmentName": "210车间",
                "materialCode": "M-004",
                "inventoryName": "零计划药品",
                "specification": "1mg×10片",
                "monthlyProductionPlan": 0,
            },
            {
                "departmentName": "302车间",
                "materialCode": "M-005",
                "inventoryName": "阿德福韦酯片",
                "specification": "10mg×28片×100瓶",
                "monthlyProductionPlan": 50,
            },
            {
                "departmentName": "302车间",
                "materialCode": "M-007",
                "inventoryName": "阿德福韦酯片",
                "specification": "20mg×10片",
                "monthlyProductionPlan": 80,
            },
            {
                "departmentName": "302车间",
                "materialCode": "M-008",
                "inventoryName": "不存在品种",
                "specification": "10mg×28片×100瓶",
                "monthlyProductionPlan": 90,
            },
            {
                "departmentName": "302车间",
                "materialCode": "M-009",
                "inventoryName": "阿司匹林肠溶片100mg",
                "specification": "10mg×28片×100瓶",
                "monthlyProductionPlan": 40,
            },
            {
                "departmentName": "303车间",
                "materialCode": "M-006",
                "inventoryName": "完全未匹配药品",
                "specification": "30mg×10片",
                "monthlyProductionPlan": 60,
            },
            {
                "departmentName": "303车间",
                "materialCode": "M-006B",
                "inventoryName": "完全未匹配药品",
                "specification": "30mg×10片",
                "monthlyProductionPlan": 70,
            },
        ]
        UploadFileItem.objects.bulk_create([
            UploadFileItem(file=task.file, **row) for row in plan_rows
        ])
        ApsArchiveItem.objects.bulk_create([
            ApsArchiveItem(
                archive=task.apsArchive,
                productName="阿司匹林肠溶片100mg",
                packageSpecification="47.5mg×14片×2板×400盒",
            ),
            ApsArchiveItem(
                archive=task.apsArchive,
                productName="阿德福韦酯片",
                packageSpecification="10mg×28片×100瓶",
            ),
        ])

        response = self.client.post(
            "/tdsms/solve/matchCheck",
            {"taskId": task.taskId},
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        result = response.data["data"]
        self.assertFalse(result["status"])
        self.assertEqual(result["page"], 1)
        self.assertEqual(result["pageSize"], 10)
        missing = {
            (record["inventoryName"], record["specification"]): record["reason"]
            for record in result["missingData"]
        }
        self.assertEqual(result["total"], 6)
        self.assertEqual(len(result["missingData"]), 6)
        self.assertEqual(
            set(missing),
            {
                ("未维护药品", "20mg×10片"),
                ("零计划药品", "1mg×10片"),
                ("阿德福韦酯片", "20mg×10片"),
                ("不存在品种", "10mg×28片×100瓶"),
                ("阿司匹林肠溶片100mg", "10mg×28片×100瓶"),
                ("完全未匹配药品", "30mg×10片"),
            },
        )
        self.assertEqual(missing[("阿德福韦酯片", "20mg×10片")], REASON_MISSING_SPEC)
        self.assertEqual(missing[("不存在品种", "10mg×28片×100瓶")], REASON_MISSING_PRODUCT)
        self.assertEqual(missing[("未维护药品", "20mg×10片")], REASON_MISSING_BOTH)
        self.assertEqual(missing[("阿司匹林肠溶片100mg", "10mg×28片×100瓶")], REASON_MISSING_BOTH)
        self.assertEqual(missing[("零计划药品", "1mg×10片")], REASON_MISSING_BOTH)
        self.assertEqual(missing[("完全未匹配药品", "30mg×10片")], REASON_MISSING_BOTH)
        self.assertNotIn(("阿司匹林肠溶片（过评）", "47.5mg×14粒×2板×400盒"), missing)
        self.assertNotIn(("阿德福韦酯片", "10mg×28片×100瓶"), missing)
        for record in result["missingData"]:
            self.assertEqual(set(record), {"inventoryName", "specification", "reason"})

        paged = self.client.post(
            "/tdsms/solve/matchCheck",
            {"taskId": task.taskId, "page": 2, "pageSize": 2},
            format="json",
        )
        self.assertEqual(paged.status_code, 200, paged.data)
        paged_data = paged.data["data"]
        self.assertFalse(paged_data["status"])
        self.assertEqual(paged_data["total"], 6)
        self.assertEqual(paged_data["page"], 2)
        self.assertEqual(paged_data["pageSize"], 2)
        self.assertEqual(len(paged_data["missingData"]), 2)
        self.assertEqual(
            paged_data["missingData"],
            result["missingData"][2:4],
        )

        start_payload = {
            "taskId": task.taskId,
            "department": "210车间",
            "scheduleMonth": "2025-06",
            "productionRules": {
                "continuousRunLimitDays": 5.5,
                "cleaningDuration": {
                    "majorCleaningDays": 0.5,
                    "minorCleaningDays": 0.25,
                    "periodicCleaningDays": 0.5,
                },
                "shiftConversion": {"naturalDays": 1, "shiftCount": 2},
            },
            "personnelCapacity": {
                "mixing": 3,
                "tableting": 2,
                "coating": 2,
                "packaging": 4,
            },
            "solverTimeLimitMinutes": 20,
        }
        with patch("api.views.solve_views.submit_solver"):
            started = self.client.post("/tdsms/solve/start", start_payload, format="json")
        self.assertEqual(started.status_code, 200, started.data)
        self.assertEqual(started.data["message"], "求解任务已创建")
        solve = SolveTask.objects.get(solveTaskId=started.data["data"]["solveTaskId"])
        self.assertEqual(solve.inputParams["unmatchedItemPolicy"], "BLOCK")
        self.assertNotIn("matchedCount", started.data["data"])
        self.assertNotIn("skippedCount", started.data["data"])
        solve.solveStatus = 2
        solve.save(update_fields=["solveStatus"])

        with patch("api.views.solve_views.submit_solver"):
            skipped = self.client.post(
                "/tdsms/solve/start",
                {**start_payload, "unmatchedItemPolicy": "SKIP"},
                format="json",
            )
        self.assertEqual(skipped.status_code, 200, skipped.data)
        solve = SolveTask.objects.get(solveTaskId=skipped.data["data"]["solveTaskId"])
        self.assertEqual(solve.inputParams["unmatchedItemPolicy"], "SKIP")
        self.assertNotIn("skippedCount", solve.inputParams)
        solve.solveStatus = 2
        solve.save(update_fields=["solveStatus"])

        with patch("api.views.solve_views.submit_solver"):
            all_missing = self.client.post(
                "/tdsms/solve/start",
                {
                    **start_payload,
                    "department": "303车间",
                    "unmatchedItemPolicy": "SKIP",
                },
                format="json",
            )
        self.assertEqual(all_missing.status_code, 200, all_missing.data)

        self.client.credentials()
        self.login("user02", "123456")
        hidden = self.client.post(
            "/tdsms/solve/matchCheck",
            {"taskId": task.taskId},
            format="json",
        )
        self.assertEqual(hidden.status_code, 404)

    def test_legacy_running_solve_without_algorithm_directory_can_be_stopped(self):
        from algorithm.adapter import stop_solver

        task = self.create_task()
        solve = SolveTask.objects.create(
            task=task,
            createdUser=self.user,
            inputParams={},
            solveStatus=1,
        )
        result = stop_solver(solve.solveTaskId)

        solve.refresh_from_db()
        self.assertEqual(result["status"], "STOPPED")
        self.assertTrue(result["orphaned"])
        self.assertEqual(solve.solveStatus, 4)
        self.assertEqual(solve.finishReason, 3)
        self.assertIsNotNone(solve.finishTime)

    def test_stopping_task_marks_stopped_immediately_and_keeps_partial_result(self):
        from algorithm.adapter import sync_solve_task

        solve = self.create_solve_record()
        run_dir = Path(settings.ALGORITHM_RUN_ROOT) / str(solve.solveTaskId)
        run_dir.mkdir(parents=True, exist_ok=True)
        visual = run_dir / "可排产结果可视化.xlsx"
        visual.unlink(missing_ok=True)
        engine = Mock()
        engine.query_solver_status.return_value = {
            "taskId": str(solve.solveTaskId),
            "status": "STOPPING",
            "startedAt": "2026-08-17T12:00:00",
            "finishedAt": None,
            "stopRequested": True,
            "resultReady": False,
            "progressMessage": "收到停止请求，正在导出当前最优方案",
            "resultFiles": {"visualBoard": str(visual)},
        }

        with patch("algorithm.adapter._engine_service", return_value=engine):
            sync_solve_task(solve, cleanup=False)

        solve.refresh_from_db()
        self.assertEqual(solve.solveStatus, 4)
        self.assertEqual(solve.finishReason, 3)
        self.assertIsNotNone(solve.finishTime)
        self.assertFalse(solve.partialResultFilePath)

        visual.write_bytes(b"result")
        engine.query_solver_status.return_value.update({
            "status": "STOPPED",
            "finishedAt": "2026-08-17T12:00:05",
            "resultReady": True,
            "progressMessage": "已停止求解，当前最优结果已生成",
        })
        with patch("algorithm.adapter._engine_service", return_value=engine):
            sync_solve_task(solve, cleanup=False)

        solve.refresh_from_db()
        self.assertEqual(solve.solveStatus, 4)
        self.assertEqual(solve.finishReason, 3)
        self.assertEqual(solve.partialResultFilePath, str(visual))

    def test_running_partial_result_is_saved_only_after_visual_file_exists(self):
        from algorithm.adapter import sync_solve_task

        solve = self.create_solve_record()
        run_dir = Path(settings.ALGORITHM_RUN_ROOT) / str(solve.solveTaskId)
        run_dir.mkdir(parents=True, exist_ok=True)
        visual = run_dir / "可排产结果可视化.xlsx"
        visual.unlink(missing_ok=True)
        engine = Mock()
        engine.query_solver_status.return_value = {
            "taskId": str(solve.solveTaskId),
            "status": "RUNNING",
            "resultReady": True,
            "resultKind": "partial",
            "resultFiles": {"visualBoard": str(visual)},
        }

        with patch("algorithm.adapter._engine_service", return_value=engine):
            sync_solve_task(solve, cleanup=False)

        solve.refresh_from_db()
        self.assertFalse(solve.partialResultFilePath)
        self.assertFalse(solve.resultFilePath)

        visual.write_bytes(b"result")
        with patch("algorithm.adapter._engine_service", return_value=engine):
            sync_solve_task(solve, cleanup=False)

        solve.refresh_from_db()
        self.assertEqual(solve.partialResultFilePath, str(visual))
        self.assertFalse(solve.resultFilePath)

        engine.query_solver_status.return_value.update({
            "status": "SUCCESS",
            "finishedAt": "2026-08-18T12:00:00",
            "resultKind": "final",
            "result": {"cp_status": "FEASIBLE"},
        })
        with patch("algorithm.adapter._engine_service", return_value=engine):
            sync_solve_task(solve, cleanup=False)

        solve.refresh_from_db()
        self.assertEqual(solve.resultFilePath, str(visual))

    def test_solve_query_flags_require_generated_result_files(self):
        solve = self.create_solve_record(
            partialResultFilePath=r"C:\missing\partial-result.xlsx",
            resultFilePath=r"C:\missing\final-result.xlsx",
        )

        with patch("api.views.solve_views.sync_solve_task", return_value=(solve, None)):
            response = self.client.get("/tdsms/solve/query", {"solveTaskId": solve.solveTaskId})

        self.assertEqual(response.status_code, 200, response.data)
        self.assertFalse(response.data["data"]["hasPartialResult"])
        self.assertFalse(response.data["data"]["hasFinalResult"])

        run_dir = Path(settings.ALGORITHM_RUN_ROOT) / str(solve.solveTaskId)
        run_dir.mkdir(parents=True, exist_ok=True)
        visual = run_dir / "可排产结果可视化.xlsx"
        visual.write_bytes(b"result")
        solve.partialResultFilePath = str(visual)
        solve.resultFilePath = None
        solve.save(update_fields=["partialResultFilePath", "resultFilePath"])

        with patch("api.views.solve_views.sync_solve_task", return_value=(solve, None)):
            response = self.client.get("/tdsms/solve/query", {"solveTaskId": solve.solveTaskId})

        self.assertTrue(response.data["data"]["hasPartialResult"])
        self.assertFalse(response.data["data"]["hasFinalResult"])

        solve.resultFilePath = str(visual)
        solve.save(update_fields=["resultFilePath"])
        with patch("api.views.solve_views.sync_solve_task", return_value=(solve, None)):
            response = self.client.get("/tdsms/solve/query", {"solveTaskId": solve.solveTaskId})

        self.assertTrue(response.data["data"]["hasPartialResult"])
        self.assertTrue(response.data["data"]["hasFinalResult"])

    def test_result_download_allows_running_task_after_partial_file_is_generated(self):
        visual = Path(self.temp_dir) / "partial-visual.xlsx"
        visual.write_bytes(b"partial-result")
        solve = self.create_solve_record(partialResultFilePath=str(visual))
        state = {"status": "RUNNING", "resultKind": "partial", "resultReady": True}

        with patch("api.views.solve_views.sync_solve_task", return_value=(solve, state)):
            response = self.client.post(
                "/tdsms/solve/result",
                {"solveTaskId": solve.solveTaskId},
                format="json",
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "可排产结果可视化.xlsx",
            unquote(response.headers["Content-Disposition"]),
        )
        self.assertEqual(
            response.headers["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(b"".join(response.streaming_content), b"partial-result")

    def test_export_returns_conflict_while_stopped_result_is_being_generated(self):
        task = self.create_task()
        solve = SolveTask.objects.create(
            task=task,
            createdUser=self.user,
            inputParams={},
            solveStatus=1,
        )
        state = {
            "status": "STOPPING",
            "finishedAt": None,
            "stopRequested": True,
            "resultReady": False,
        }
        with patch("api.views.solve_views.sync_solve_task", return_value=(solve, state)):
            response = self.client.post(
                "/tdsms/solve/result",
                {"solveTaskId": solve.solveTaskId},
                format="json",
            )

        self.assertEqual(response.status_code, 409)
        self.assertFalse(response.data["success"])
        self.assertIn("结果文件尚未生成", response.data["message"])

    @patch("api.views.solve_views.submit_solver")
    def test_start_stops_current_user_running_solve_then_starts_new(
        self,
        submit_solver_mock,
    ):
        self.login()
        archive = ApsArchive.objects.create(archiveName="测试档案", createdBy=self.user)
        ApsArchiveItem.objects.create(
            archive=archive,
            productName="阿德福韦酯片",
            packageSpecification="10mg×28片×100瓶",
        )
        upload = UploadFile.objects.create(
            originalName="plan.xlsx",
            fileName="plan.xlsx",
            filePath="plan.xlsx",
            uploadUser=self.user,
            parseStatus=1,
        )
        UploadFileItem.objects.create(
            file=upload,
            departmentName="210车间",
            materialCode="M-001",
            inventoryName="阿德福韦酯片",
            specification="10mg×28片×100瓶",
            monthlyProductionPlan=100,
        )
        task = TaskImportRecord.objects.create(
            apsArchive=archive,
            file=upload,
            importStatus=1,
            createdBy=self.user,
        )

        def payload(task_id):
            return {
                "taskId": task_id,
                "department": "210车间",
                "scheduleMonth": "2025-06",
                "unmatchedItemPolicy": "SKIP",
                "productionRules": {
                    "continuousRunLimitDays": 5.5,
                    "cleaningDuration": {
                        "majorCleaningDays": 0.5,
                        "minorCleaningDays": 0.25,
                        "periodicCleaningDays": 0.5,
                    },
                    "shiftConversion": {"naturalDays": 1, "shiftCount": 2},
                },
                "personnelCapacity": {
                    "mixing": 3,
                    "tableting": 2,
                    "coating": 2,
                    "packaging": 4,
                },
                "solverTimeLimitMinutes": 20,
            }

        first = self.client.post("/tdsms/solve/start", payload(task.taskId), format="json")
        self.assertEqual(first.status_code, 200, first.data)
        first_solve_id = first.data["data"]["solveTaskId"]

        another_task = TaskImportRecord.objects.create(
            sourceType=2,
            sourceTask=task,
            apsArchive=task.apsArchive,
            file=task.file,
            importStatus=1,
            createdBy=self.user,
        )
        replaced = self.client.post(
            "/tdsms/solve/start",
            payload(another_task.taskId),
            format="json",
        )
        self.assertEqual(replaced.status_code, 200, replaced.data)
        second_solve_id = replaced.data["data"]["solveTaskId"]
        self.assertNotEqual(second_solve_id, first_solve_id)
        self.assertEqual(SolveTask.objects.get(solveTaskId=first_solve_id).solveStatus, 4)
        self.assertEqual(SolveTask.objects.get(solveTaskId=second_solve_id).solveStatus, 0)

        other_archive = ApsArchive.objects.create(archiveName="对方档案", createdBy=self.other)
        ApsArchiveItem.objects.create(
            archive=other_archive,
            productName="阿德福韦酯片",
            packageSpecification="10mg×28片×100瓶",
        )
        other_upload = UploadFile.objects.create(
            originalName="plan.xlsx",
            fileName="plan2.xlsx",
            filePath="plan2.xlsx",
            uploadUser=self.other,
            parseStatus=1,
        )
        UploadFileItem.objects.create(
            file=other_upload,
            departmentName="210车间",
            materialCode="M-001",
            inventoryName="阿德福韦酯片",
            specification="10mg×28片×100瓶",
            monthlyProductionPlan=100,
        )
        other_task = TaskImportRecord.objects.create(
            apsArchive=other_archive,
            file=other_upload,
            importStatus=1,
            createdBy=self.other,
        )
        self.client.credentials()
        self.login("user02", "123456")
        other_started = self.client.post(
            "/tdsms/solve/start",
            payload(other_task.taskId),
            format="json",
        )
        self.assertEqual(other_started.status_code, 200, other_started.data)
        self.assertEqual(SolveTask.objects.get(solveTaskId=second_solve_id).solveStatus, 0)
        self.assertEqual(submit_solver_mock.call_count, 3)
