from datetime import datetime
from pathlib import Path
from unittest import mock
import tempfile
import unittest

import pandas as pd
from openpyxl import Workbook


class DynamicShiftConversionTests(unittest.TestCase):
    def test_sunday_conversion_skips_three_shifts(self):
        from postprocess_sunday_calendar import working_to_calendar_shift

        start_date = datetime(2026, 7, 4)  # Saturday

        self.assertEqual(working_to_calendar_shift(3, start_date, shifts_per_day=3), 6)
        self.assertEqual(working_to_calendar_shift(4, start_date, shifts_per_day=3), 7)

    def test_weekly_split_uses_three_shifts_per_day(self):
        from output_weekly_template import _split_task_by_week

        start_date = datetime(2026, 7, 6)  # Monday
        rows = _split_task_by_week(0, 3, 1, start_date, shifts_per_day=3)

        slots = rows[datetime(2026, 7, 5)]
        self.assertEqual(len(slots), 21)
        self.assertEqual(slots[3:6], [1.0, 1.0, 1.0])
        self.assertEqual(sum(slots), 3.0)

    def test_weekly_split_recomputes_quantities_for_two_three_and_five_shifts(self):
        from output_weekly_template import _split_task_by_week

        start_date = datetime(2026, 7, 6)  # Monday

        expected = {
            2: (14, [75.0, 100.0, 75.0], 250.0),
            3: (21, [75.0, 100.0, 75.0], 250.0),
            5: (35, [75.0, 100.0, 75.0], 250.0),
        }

        for shifts_per_day, (slot_count, non_empty, total) in expected.items():
            with self.subTest(shifts_per_day=shifts_per_day):
                rows = _split_task_by_week(
                    4.25,
                    6.75,
                    100,
                    start_date,
                    shifts_per_day=shifts_per_day,
                )

                slots = rows[datetime(2026, 7, 5)]
                values = [value for value in slots if value]
                self.assertEqual(len(slots), slot_count)
                self.assertEqual(values, non_empty)
                self.assertEqual(sum(slots), total)

    def test_weekly_split_places_five_shift_boundary_in_actual_slots(self):
        from output_weekly_template import _split_task_by_week

        start_date = datetime(2026, 7, 6)  # Monday
        rows = _split_task_by_week(4.25, 6.75, 100, start_date, shifts_per_day=5)

        slots = rows[datetime(2026, 7, 5)]
        self.assertEqual(slots[9], 75.0)  # Monday5
        self.assertEqual(slots[10], 100.0)  # Tuesday1
        self.assertEqual(slots[11], 75.0)  # Tuesday2

    def test_weekly_header_supports_five_shifts(self):
        from output_weekly_template import _build_header

        header = _build_header(5)

        self.assertEqual(len(header), 44)
        self.assertEqual(header[6:41], [f"{day}{shift}" for day in (
            "周日", "周一", "周二", "周三", "周四", "周五", "周六"
        ) for shift in range(1, 6)])
        self.assertIn("周六5", header)

    def test_sunday_conversion_skips_five_shifts(self):
        from postprocess_sunday_calendar import working_to_calendar_shift

        start_date = datetime(2026, 7, 4)  # Saturday

        self.assertEqual(working_to_calendar_shift(5, start_date, shifts_per_day=5), 10)
        self.assertEqual(working_to_calendar_shift(6, start_date, shifts_per_day=5), 11)

    def test_split_rejects_non_positive_or_fractional_shift_counts(self):
        from output_weekly_template import _split_task_by_week

        start_date = datetime(2026, 7, 6)

        for shifts_per_day in (0, -1, 2.5):
            with self.subTest(shifts_per_day=shifts_per_day):
                with self.assertRaises(ValueError):
                    _split_task_by_week(0, 1, 100, start_date, shifts_per_day=shifts_per_day)

    def test_write_block_places_summary_columns_after_dynamic_shift_columns(self):
        from output_weekly_template import _write_block

        for shifts_per_day in (2, 3, 5):
            with self.subTest(shifts_per_day=shifts_per_day):
                wb = Workbook()
                ws = wb.active
                total_col = 7 + 7 * shifts_per_day
                metadata = {
                    "drug spec": {
                        "产品代码": "P001",
                        "品名": "Drug",
                        "包装规格": "Box",
                        "销售计划量": 123,
                    }
                }
                slots = [0.0] * (7 * shifts_per_day)
                slots[-1] = 12.5

                _write_block(
                    ws,
                    1,
                    datetime(2026, 7, 5),
                    1,
                    [("Line A", "drug spec", slots)],
                    metadata,
                    department="210车间",
                    shifts_per_day=shifts_per_day,
                )

                self.assertEqual(ws.cell(3, total_col).value, "生产计划量")
                self.assertEqual(ws.cell(3, total_col + 1).value, "销售计划量")
                self.assertEqual(ws.cell(3, total_col + 2).value, "计划说明")
                self.assertEqual(ws.cell(4, total_col).value, 12.5)

    def test_collect_visual_rows_passes_shifts_per_day_to_input_builder(self):
        from output_weekly_template import _collect_visual_rows

        with mock.patch("output_weekly_template.build_schedule_inputs") as builder:
            builder.return_value = (
                [], [], [], [], [], {}, {}, {}, {}, {}, {}, {}, {}, {},
                {}, {}, {}, {}, None,
            )
            with mock.patch("output_weekly_template._build_stage_rates", return_value={1: {}, 2: {}, 3: {}, 4: {}}), \
                    mock.patch("output_weekly_template._build_item_metadata", return_value={}), \
                    mock.patch("pandas.read_excel") as read_excel:
                read_excel.side_effect = [
                    mock.Mock(iterrows=lambda: iter([])),
                    mock.Mock(iterrows=lambda: iter([])),
                ]

                _collect_visual_rows(
                    "result.xlsx",
                    "demo.xlsx",
                    "aps.xlsx",
                    datetime(2026, 7, 1),
                    shifts_per_day=5,
                )

        self.assertEqual(builder.call_args.kwargs["shifts_per_day"], 5)

    def test_output_template_cli_passes_shifts_per_day_to_generator(self):
        import output_weekly_template

        with mock.patch.object(
            output_weekly_template,
            "generate_template_schedule_board",
            return_value="可排产结果可视化.xlsx",
        ) as generator, mock.patch(
            "sys.argv",
            [
                "output_weekly_template.py",
                "--shifts-per-day",
                "3",
                "--department",
                "210车间",
            ],
        ):
            output_weekly_template.main()

        self.assertEqual(generator.call_args.kwargs["shifts_per_day"], 3)
        self.assertEqual(generator.call_args.kwargs["department"], "210车间")

    def test_board_groups_by_stage_and_sorts_rows_by_start_time(self):
        import output_weekly_template
        from output_weekly_template import generate_template_schedule_board

        week1 = datetime(2026, 7, 5)
        week2 = datetime(2026, 7, 12)
        slots_late = [0.0] * 14
        slots_late[5] = 10.0
        slots_early = [0.0] * 14
        slots_early[1] = 8.0
        slots_week2 = [0.0] * 14
        slots_week2[0] = 6.0

        rows = {
            (week2, 1, "B线", "item-b"): slots_week2,
            (week1, 1, "Z线", "item-late"): slots_late,
            (week1, 1, "A线", "item-early"): slots_early,
            (week1, 2, "压片1", "item-press"): slots_early,
        }
        metadata = {
            "item-b": {"产品代码": "B", "品名": "B药", "包装规格": "盒", "销售计划量": 1},
            "item-late": {"产品代码": "L", "品名": "L药", "包装规格": "盒", "销售计划量": 1},
            "item-early": {"产品代码": "E", "品名": "E药", "包装规格": "盒", "销售计划量": 1},
            "item-press": {"产品代码": "P", "品名": "P药", "包装规格": "盒", "销售计划量": 1},
        }

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "board.xlsx"
            with mock.patch.object(
                output_weekly_template,
                "_collect_visual_rows",
                return_value=(rows, metadata),
            ), mock.patch.object(
                output_weekly_template,
                "get_mix_spec",
                return_value="规格",
            ):
                generate_template_schedule_board(
                    "result.xlsx",
                    "demo.xlsx",
                    "aps.xlsx",
                    output,
                    start_date=datetime(2026, 7, 1),
                )

            from openpyxl import load_workbook

            ws = load_workbook(output).active
            texts = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value]
            stage_headers = [t for t in texts if "·" in str(t)]
            self.assertEqual(
                stage_headers,
                ["210车间 · 配料计划", "210车间 · 压片计划"],
            )
            mix_header_row = next(
                r for r in range(1, ws.max_row + 1)
                if ws.cell(r, 1).value == "210车间 · 配料计划"
            )
            press_header_row = next(
                r for r in range(1, ws.max_row + 1)
                if ws.cell(r, 1).value == "210车间 · 压片计划"
            )
            self.assertLess(mix_header_row, press_header_row)

            # 配料板块内：先 week1（early 再 late），再 week2
            devices = [
                ws.cell(r, 1).value
                for r in range(mix_header_row + 1, press_header_row)
                if ws.cell(r, 1).value in {"A线", "Z线", "B线"}
            ]
            self.assertEqual(devices, ["A线", "Z线", "B线"])

    def test_sunday_postprocess_passes_shifts_per_day_to_input_builder(self):
        from postprocess_sunday_calendar import postprocess_detail_for_sundays

        with mock.patch("postprocess_sunday_calendar.build_schedule_inputs") as builder:
            builder.return_value = (
                [], [], [], [], [], {}, {}, {}, {}, {}, {}, {}, {}, {},
                {}, {}, {}, {}, None,
            )
            batch_df = pd.DataFrame(
                columns=["药品规格", "批次号", "配料开工(班时)", "批次周期上限(班时)"]
            )
            pack_df = pd.DataFrame(
                columns=["药品规格", "总批次", "包装开工(班时)", "包装完工(班时)"]
            )
            clear_df = pd.DataFrame(
                columns=[
                    "工序",
                    "前任务药品规格",
                    "前任务批次",
                    "前任务结束(班时)",
                    "后任务药品规格",
                    "后任务批次",
                    "后任务开始(班时)",
                    "理论清场时长",
                ]
            )
            with mock.patch("pandas.read_excel", side_effect=[batch_df, pack_df, clear_df]), \
                    mock.patch("pandas.ExcelWriter") as writer, \
                    mock.patch("pandas.DataFrame.to_excel"):
                writer.return_value.__enter__.return_value = mock.Mock()

                postprocess_detail_for_sundays(
                    "input.xlsx",
                    "output.xlsx",
                    "demo.xlsx",
                    "aps.xlsx",
                    datetime(2026, 7, 1),
                    shifts_per_day=5,
                )

        self.assertEqual(builder.call_args.kwargs["shifts_per_day"], 5)


if __name__ == "__main__":
    unittest.main()
