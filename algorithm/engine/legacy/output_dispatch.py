import pandas as pd
import numpy as np
import math
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment


def auto_expand_excel_cells(filename):
    """
    美化增强版：
    1. 强制开启自动换行。
    2. 显著提升行高系数（设为 25），增加上下留白。
    3. 居中对齐优化。
    """
    print(f"🎨 正在应用【宽行模式】美化处理...")
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        start_row = 4
        start_col = 3

        for row in range(start_row, ws.max_row + 1):
            max_line_count = 1
            for col in range(start_col, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    # 设置垂直居中，开启自动换行
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                    line_count = str(cell.value).count('\n') + 1
                    if line_count > max_line_count:
                        max_line_count = line_count

            # 将行高系数设为 45
            ws.row_dimensions[row].height = max_line_count * 45

        # 适当拓宽列宽，给长规格名称更多空间
        for col in range(start_col, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[col_letter].width = 45

        wb.save(filename)
        print(f"✅ 格式美化完成，单元格已全部自动展开！请查看文件: {filename}")
    except Exception as e:
        print(f"❌ 美化过程出错: {e}")


def generate_factory_dispatch_bill(schedule_file, aps_file, output_name):
    print("=========================================")
    print("🚀 启动派工单生成引擎 (包含高精度小数与自动展开排版)")
    print("=========================================")

    # 1. 加载排产结果
    try:
        df_batch_res = pd.read_excel(schedule_file, sheet_name='批次工序计划')
        df_pack_res = pd.read_excel(schedule_file, sheet_name='包装及摘要')
        try:
            df_clear_res = pd.read_excel(schedule_file, sheet_name='设备清场明细')
        except Exception:
            df_clear_res = pd.DataFrame()
    except Exception as e:
        print(f"❌ 读取排产结果失败: {e}")
        return False

    # 2. 严谨解析 APS 基础工艺数据
    df_aps_raw = pd.read_excel(aps_file, sheet_name='APS生产信息统计')
    columns_l1 = df_aps_raw.iloc[2].fillna(method='ffill').tolist()
    columns_l2 = df_aps_raw.iloc[3].tolist()
    cols = []
    for i, (c1, c2) in enumerate(zip(columns_l1, columns_l2)):
        c1_str = str(c1).strip() if pd.notna(c1) else ""
        c2_str = str(c2).strip() if pd.notna(c2) else ""
        cols.append(f"{c1_str}_{c2_str}" if c1_str and c2_str else (c1_str or c2_str or f"Un_{i}"))

    df_aps = df_aps_raw.iloc[4:].copy()
    df_aps.columns = cols

    df_aps['品种'] = df_aps['品种'].fillna(method='ffill').astype(str).str.strip()
    df_aps['包装规格'] = df_aps['包装规格'].astype(str).str.strip()

    fill_cols = [
        '配料_批量（万片/粒）', '配料_班产量（万片）',
        '压片_班产量', '包衣_班产量', '分装/铝塑_班产量（万片）'
    ]
    for col in fill_cols:
        if col in df_aps.columns:
            df_aps[col] = df_aps[col].replace(r'^\s*$', np.nan, regex=True)
            df_aps[col] = df_aps.groupby('品种')[col].ffill()

    def sf(val):
        try:
            return float(val)
        except:
            return 0.0

    aps_data = {}
    for _, row in df_aps.iterrows():
        key = f"{str(row['品种'])} {str(row['包装规格'])}"
        b_size = sf(row.get('配料_批量（万片/粒）', 0))
        aps_data[key] = {
            'batch_size': b_size,
            'mix_p': (b_size / sf(row.get('配料_班产量（万片）', 0))) if sf(row.get('配料_班产量（万片）', 0)) > 0 else 0,
            'tab_p': (b_size / sf(row.get('压片_班产量', 0))) if sf(row.get('压片_班产量', 0)) > 0 else 0,
            'coat_p': (b_size / sf(row.get('包衣_班产量', 0))) if sf(row.get('包衣_班产量', 0)) > 0 else 0,
            'pack_p': (b_size / sf(row.get('分装/铝塑_班产量（万片）', 0))) if sf(
                row.get('分装/铝塑_班产量（万片）', 0)) > 0 else 0,
        }

        # 3. 动态时间轴计算 (增加星期维度)
        clear_max_t = 0
        if not df_clear_res.empty and '清场结束(班时)' in df_clear_res.columns:
            clear_max_t = pd.to_numeric(df_clear_res['清场结束(班时)'], errors='coerce').max() or 0
        max_t = max(pd.to_numeric(df_batch_res['配料开工(班时)'], errors='coerce').max() or 0,
                    df_pack_res['包装完工(班时)'].max() if '包装完工(班时)' in df_pack_res.columns else 0,
                    clear_max_t)
        total_days = int(math.ceil((max_t + 1) / 2))
        start_date = datetime(2026, 7, 1)

        # 🌟 星期的中文映射
        weekday_map = {0: '周一', 1: '周二', 2: '周三', 3: '周四', 4: '周五', 5: '周六', 6: '周日'}

        dates_list = []
        wday_list = []
        for i in range(total_days):
            curr_date = start_date + timedelta(days=i)
            dates_list.append(curr_date.strftime('%m/%d'))
            wday_list.append(weekday_map[curr_date.weekday()])  # 计算出这一天是星期几

        # 🌟 生成三层表头：日期 -> 星期 -> 班次
        col_tuples = [(dates_list[i], wday_list[i], s) for i in range(total_days) for s in ['早班', '晚班']]
        col_index = pd.MultiIndex.from_tuples(col_tuples, names=['日期', '星期', '班次'])

    # 4. 构建设备行索引
    def get_macs(df, col):
        return [str(m).strip() for m in df[col].unique() if
                str(m).strip() not in ['nan', '-', '未分配', '无', '']]

    row_tuples = [('1. 配料', m) for m in get_macs(df_batch_res, '配料设备')] + \
                 [('2. 压片', m) for m in get_macs(df_batch_res, '压片设备')] + \
                 [('3. 包衣', m) for m in get_macs(df_batch_res, '包衣设备')] + \
                 [('4. 包装', m) for m in get_macs(df_pack_res, '分装铝塑设备')]
    calendar_df = pd.DataFrame("", index=pd.MultiIndex.from_tuples(row_tuples), columns=col_index)

    # 5. 聚合产量 (精准切分 + 记录开工时间)
    temp_storage = {}

    def allocate(stage, mac, drug_key, start, dur, total_qty):
        if dur <= 0 or total_qty <= 0: return
        end = start + dur
        for s_idx in range(int(start), int(math.ceil(end))):
            overlap = max(0, min(end, s_idx + 1) - max(start, s_idx))
            if overlap > 0.001 and (s_idx // 2) < total_days:
                val = (overlap / dur) * total_qty
                # 🌟 写入时增加星期维度 wday_list[s_idx // 2]
                key = (stage, mac, dates_list[s_idx // 2], wday_list[s_idx // 2], '早班' if s_idx % 2 == 0 else '晚班',
                       drug_key)

                # 计算在该班次内的实际开工时间，用于后续文本上下排序
                actual_start_in_shift = max(start, s_idx)

                if key not in temp_storage:
                    temp_storage[key] = {'qty': 0.0, 'start_time': actual_start_in_shift}

                temp_storage[key]['qty'] += val
                # 如果同一药品在该班次有多批，保留最早的开工时间
                temp_storage[key]['start_time'] = min(temp_storage[key]['start_time'], actual_start_in_shift)

    def allocate_event(stage, mac, label, start, dur):
        if dur <= 0:
            return
        end = start + dur
        for s_idx in range(int(start), int(math.ceil(end))):
            overlap = max(0, min(end, s_idx + 1) - max(start, s_idx))
            if overlap > 0.001 and (s_idx // 2) < total_days:
                key = (
                    stage, mac, dates_list[s_idx // 2], wday_list[s_idx // 2],
                    '早班' if s_idx % 2 == 0 else '晚班', label
                )
                actual_start_in_shift = max(start, s_idx)
                if key not in temp_storage:
                    temp_storage[key] = {'qty': 0.0, 'start_time': actual_start_in_shift}
                temp_storage[key]['qty'] += overlap
                temp_storage[key]['start_time'] = min(temp_storage[key]['start_time'], actual_start_in_shift)

    # 6. 遍历分配
    configs = [('配料开工(班时)', '配料设备', '1. 配料', 'mix_p'), ('压片开工(班时)', '压片设备', '2. 压片', 'tab_p'),
               ('包衣开工(班时)', '包衣设备', '3. 包衣', 'coat_p')]
    for _, row in df_batch_res.iterrows():
        drug = str(row['药品规格'])
        if drug in aps_data:
            for t_col, m_col, stg, p_k in configs:
                mac, start = str(row[m_col]).strip(), row[t_col]
                if mac in calendar_df.index.get_level_values(1) and pd.notna(start) and str(start) != '无需包衣':
                    allocate(stg, mac, drug, float(start), aps_data[drug][p_k], aps_data[drug]['batch_size'])

    for _, row in df_pack_res.iterrows():
        drug = str(row['药品规格'])
        mac = str(row['分装铝塑设备']).strip()
        if drug in aps_data and mac in calendar_df.index.get_level_values(1):
            s, e = float(row['包装开工(班时)']), float(row['包装完工(班时)'])
            allocate('4. 包装', mac, drug, s, e - s, row['总批次'] * aps_data[drug]['batch_size'])

    if not df_clear_res.empty:
        for _, row in df_clear_res.iterrows():
            clear_dur = float(row.get('理论清场时长', 0) or 0)
            if clear_dur <= 0:
                continue
            stage = str(row.get('工序', '')).strip()
            mac = str(row.get('设备', '')).strip()
            if mac not in calendar_df.index.get_level_values(1):
                continue
            clear_type = str(row.get('清场类型', '清场')).strip()
            start = float(row.get('清场开始(班时)', 0))
            end = float(row.get('清场结束(班时)', start + clear_dur))
            allocate_event(stage, mac, clear_type, start, end - start)

    # 7. 安全写入单元格 (带有时间轴动态排序)
    cell_buffer = {}

    for (stage, mac, d, w, s, full_name), data in temp_storage.items():
        qty = data['qty']
        start_time = data['start_time']

        if qty < 0.01:
            continue

        row_key, col_key = (stage, mac), (d, w, s)
        content = f"{full_name} {qty:.2f}"

        if (row_key, col_key) not in cell_buffer:
            cell_buffer[(row_key, col_key)] = []

        # 将时间戳和文本作为一个整体存入列表，方便后续比对
        cell_buffer[(row_key, col_key)].append((start_time, content))

    for (row_key, col_key), items in cell_buffer.items():
        # 核心逻辑：根据 start_time（x[0]）从小到大升序排序
        items.sort(key=lambda x: x[0])

        # 提取排序后的文字（x[1]）并用换行符拼接
        lines = [x[1] for x in items]
        calendar_df.at[row_key, col_key] = "\n".join(lines)

    # 保存基础 Excel
    calendar_df.to_excel(output_name)
    print(f"✅ 基础排产表已成功写入: {output_name}")

    # 🌟 关键点：生成完毕后直接调用美化函数
    auto_expand_excel_cells(output_name)
    return True


if __name__ == "__main__":
    generate_factory_dispatch_bill("排产结果明细_7月_hybrid.xlsx", "./输入/副本APS排产信息-4.30.xlsx",
                                 "210车间派工单_7月_hybrid.xlsx")
