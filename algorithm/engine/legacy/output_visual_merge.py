import openpyxl
import re
from openpyxl.styles import Alignment, PatternFill


def merge_continuous_tasks(input_file, output_file):
    print("=========================================")
    print(f"🎨 启动 Excel 横向智能合并 & 自动上色引擎")
    print("=========================================")
    print(f"正在读取文件: {input_file} ...")

    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
    except Exception as e:
        print(f"❌ 读取文件失败: {e}")
        return

    # 锁定数据区域
    start_row = 1
    start_col = 3
    merged_count = 0

    # 🌟 预设一套高颜值的浅色（马卡龙/护眼）调色板 (Hex格式)
    PALETTE = [
        "E6E6FA",  # 淡紫色 (Lavender)
        "FFB6C1",  # 浅粉色 (LightPink)
        "98FB98",  # 苍绿色 (PaleGreen)
        "FFE4B5",  # 鹿皮色 (Moccasin)
        "ADD8E6",  # 淡蓝色 (LightBlue)
        "F0E68C",  # 卡其色 (Khaki)
        "E0FFFF",  # 亮青色 (LightCyan)
        "FFDAB9",  # 肉色 (PeachPuff)
        "D8BFD8",  # 蓟色 (Thistle)
        "B0E0E6",  # 粉蓝色 (PowderBlue)
        "FFC0CB",  # 粉红色 (Pink)
        "87CEFA",  # 天蓝色 (LightSkyBlue)
        "F5DEB3",  # 小麦色 (Wheat)
    ]

    # 用于记录每种药品规格对应的颜色
    drug_color_map = {}

    # 逐行扫描
    for row in range(start_row, ws.max_row + 1):
        current_drug = None
        current_start_col = None
        accumulated_qty = 0.0

        # 多循环一次以强制触发行末结算
        for col in range(start_col, ws.max_column + 2):
            cell = ws.cell(row=row, column=col) if col <= ws.max_column else None
            val = cell.value if cell else None

            is_pure = False
            drug_name = None
            qty = 0.0

            # 1. 单元格内容解析
            if val:
                lines = str(val).strip().split('\n')
                if len(lines) == 1 and lines[0].strip():
                    # match = re.search(r'^(.*)\s+(\d+)$', lines[0].strip())
                    match = re.search(r'^(.*)\s+(\d+(?:\.\d+)?)$', lines[0].strip())
                    if match:
                        is_pure = True
                        drug_name = match.group(1).strip()
                        # qty = int(match.group(2))
                        qty = float(match.group(2))
                        # 🌟 自动分配颜色：如果是个新药品，给它按顺序分配一个颜色
                        if drug_name not in drug_color_map:
                            color_hex = "D9D9D9" if "清场" in drug_name else PALETTE[len(drug_color_map) % len(PALETTE)]
                            drug_color_map[drug_name] = PatternFill(start_color=color_hex, end_color=color_hex,
                                                                    fill_type="solid")

            # 2. 合并与上色逻辑判定
            if is_pure:
                if current_drug == drug_name:
                    accumulated_qty += qty
                else:
                    # 发现不同药品，结算前一个区块
                    if current_drug is not None:
                        # 注意：这里改成了 >=，这意味着哪怕只有一个格子没有被合并，只要它是纯净的，也会被上色和居中
                        if (col - 1) >= current_start_col:
                            _execute_merge_and_color(
                                ws, row, current_start_col, col - 1,
                                current_drug, accumulated_qty, drug_color_map[current_drug]
                            )
                            merged_count += 1

                    # 开启新追踪
                    current_drug = drug_name
                    current_start_col = col
                    accumulated_qty = qty
            else:
                # 遇到空白或混合格（交接班），结算前一个区块
                if current_drug is not None:
                    if (col - 1) >= current_start_col:
                        _execute_merge_and_color(
                            ws, row, current_start_col, col - 1,
                            current_drug, accumulated_qty, drug_color_map[current_drug]
                        )
                        merged_count += 1

                # 重置追踪器
                current_drug = None
                current_start_col = None
                accumulated_qty = 0.0

    # ==========================================
    # 🌟 新增：在合并操作完成后，强制重新刷一遍行高与自动换行
    # ==========================================
    for r in range(start_row, ws.max_row + 1):
        max_line_count = 1
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                # 重新确保合并后的所有单元格都开启了【自动换行】与垂直居中
                align = cell.alignment
                horz = align.horizontal if align else 'center'
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal=horz)

                # 统计这个格子里有几行字
                line_count = str(cell.value).count('\n') + 1
                if line_count > max_line_count:
                    max_line_count = line_count

        # 🌟 恢复你设定的 45 宽行高系数
        ws.row_dimensions[r].height = max_line_count * 45

    wb.save(output_file)
    print(f"✅ 处理完成！共执行了 {merged_count} 次区块合并与上色，并已恢复行高 45。")
    print(f"🎉 请打开 【{output_file}】 查看车间可视化看板。")


def _execute_merge_and_color(ws, row, start_col, end_col, drug_name, total_qty, fill_color):
    """
    执行单元格合并、居中并填充背景色
    """
    # 1. 写入汇总数字
    # ws.cell(row=row, column=start_col).value = f"{drug_name} {total_qty}"
    ws.cell(row=row, column=start_col).value = f"{drug_name} {total_qty:.2f}"
    # 2. 清空尾部格子避免冲突
    for c in range(start_col + 1, end_col + 1):
        ws.cell(row=row, column=c).value = None

    # 3. 跨列合并 (如果跨度大于1才执行 merge 操作)
    if end_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

    # 4. 设置居中与背景色
    target_cell = ws.cell(row=row, column=start_col)
    target_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    target_cell.fill = fill_color  # 🌟 注入背景颜色


if __name__ == "__main__":
    # 🌟 修改为读取 7天堆叠版 的数据，并输出最终高亮合并版
    INPUT_FILE = "210车间排产看板_7月_hybrid_7天堆叠版.xlsx"
    OUTPUT_FILE = "210车间排产看板_7月_hybrid_可视化看板.xlsx"
    merge_continuous_tasks(INPUT_FILE, OUTPUT_FILE)
