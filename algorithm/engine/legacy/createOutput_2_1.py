import openpyxl
import re
from openpyxl.styles import Alignment, PatternFill


def merge_continuous_tasks(input_file, output_file):
    print("=========================================")
    print(f"🎨 启动 Excel 横向智能合并 & 自动上色引擎")
    print("=========================================")
    print(f"正在读取文件: {input_file} ...")

    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
    except Exception as e:
        print(f"❌ 读取文件失败: {e}")
        return

    # 锁定数据区域
    start_row = 1
    start_col = 4  # 🌟 真正的排产产量数据自第 4 列 (D列) 开始
    merged_count = 0

    # 🌟 预设一套高颜值的浅色（马卡龙/护眼）调色板 (Hex格式)
    PALETTE = [
        "E6E6FA", "FFB6C1", "98FB98", "FFE4B5", "ADD8E6",
        "F0E68C", "E0FFFF", "FFDAB9", "D8BFD8", "B0E0E6",
        "FFC0CB", "87CEFA", "F5DEB3"
    ]

    # 用于记录每种药品规格对应的颜色
    drug_color_map = {}

    # 逐行扫描
    for row in range(start_row, ws.max_row + 1):
        current_drug = None
        current_start_col = None
        accumulated_qty = 0.0

        # 多循环一次以强制触发行末结算
        for col in range(start_col, ws.max_column + 2):
            cell = ws.cell(row=row, column=col) if col <= ws.max_column else None
            val = cell.value if cell else None

            is_pure = False
            drug_name = None
            qty = 0.0

            # 1. 单元格内容解析：现在单元格中只有纯数字产量
            if val and col <= ws.max_column:
                try:
                    qty = float(val)
                    is_pure = True
                    # 🌟 核心升级：直接从当前行的第 3 列（C列）获取药品规格名称
                    drug_name = str(ws.cell(row=row, column=3).value).strip()

                    if drug_name not in drug_color_map:
                        color_hex = PALETTE[len(drug_color_map) % len(PALETTE)]
                        drug_color_map[drug_name] = PatternFill(start_color=color_hex, end_color=color_hex,
                                                                fill_type="solid")
                except ValueError:
                    is_pure = False

            # 2. 合并与上色逻辑判定
            if is_pure:
                if current_drug == drug_name:
                    accumulated_qty += qty
                else:
                    # 发现不同药品，结算前一个区块
                    if current_drug is not None:
                        if (col - 1) >= current_start_col:
                            _execute_merge_and_color(
                                ws, row, current_start_col, col - 1,
                                current_drug, accumulated_qty, drug_color_map[current_drug]
                            )
                            merged_count += 1

                    # 开启新追踪
                    current_drug = drug_name
                    current_start_col = col
                    accumulated_qty = qty
            else:
                # 遇到空白或表头格，结算前一个区块
                if current_drug is not None:
                    if (col - 1) >= current_start_col:
                        _execute_merge_and_color(
                            ws, row, current_start_col, col - 1,
                            current_drug, accumulated_qty, drug_color_map[current_drug]
                        )
                        merged_count += 1
                current_drug = None
                current_start_col = None
                accumulated_qty = 0.0

    # ==========================================
    # 🌟 恢复你设定的 45 宽行高系数，且对齐方式自适应
    # ==========================================
    for r in range(start_row, ws.max_row + 1):
        max_line_count = 1
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                if c <= 3:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                else:
                    align = cell.alignment
                    horz = align.horizontal if align else 'right'
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal=horz)

                line_count = str(cell.value).count('\n') + 1
                if line_count > max_line_count:
                    max_line_count = line_count

        ws.row_dimensions[r].height = max_line_count * 45

    wb.save(output_file)
    print(f"✅ 处理完成！共执行了 {merged_count} 次区块合并与上色，并已恢复行高 45。")
    print(f"🎉 请打开 【{output_file}】 查看车间可视化看板。")


def _execute_merge_and_color(ws, row, start_col, end_col, drug_name, total_qty, fill_color):
    """
    执行单元格合并、居中并填充背景色
    """
    # 1. 写入汇总数字：单元格内仅写入合并后的纯数字总产量
    ws.cell(row=row, column=start_col).value = f"{total_qty:.2f}"

    # 2. 清空尾部格子避免冲突
    for c in range(start_col + 1, end_col + 1):
        ws.cell(row=row, column=c).value = None

    # 3. 跨列合并 (如果跨度大于1才执行 merge 操作)
    if end_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

    # 4. 设置居中与背景色
    target_cell = ws.cell(row=row, column=start_col)
    target_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    target_cell.fill = fill_color  # 🌟 注入背景颜色


if __name__ == "__main__":
    # 🌟 修改为读取 7天堆叠版 的数据，并输出最终高亮合并版
    INPUT_FILE = "210车间排产看板_7月7天堆叠版_1.xlsx"
    OUTPUT_FILE = "210车间排产看板_7月7天堆叠版(高亮合并终极版)_1.xlsx"
    merge_continuous_tasks(INPUT_FILE, OUTPUT_FILE)