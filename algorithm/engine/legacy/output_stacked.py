import pandas as pd
import openpyxl
from openpyxl.styles import Alignment


def convert_to_stacked_format(input_filename, output_filename, days_per_block=7):
    print("=========================================")
    print("✂️ 启动看板切片与堆叠引擎")
    print("=========================================")
    print(f"📥 正在读取原始长横轴排产表: {input_filename} ...")

    try:
        # 读取原始 Excel：指定前 3 行是表头（日期、星期、班次），前 2 列是索引（工序、设备）
        df = pd.read_excel(input_filename, header=[0, 1, 2], index_col=[0, 1])
    except Exception as e:
        print(f"❌ 读取文件失败，请确保 {input_filename} 存在且格式正确。\n错误信息: {e}")
        return

    # 每天 2 个班次，计算每块截取多少列
    cols_per_block = days_per_block * 2
    total_cols = len(df.columns)

    print(f"🔄 准备按 {days_per_block} 天（{cols_per_block} 列）为一周期进行切片堆叠...")

    # 使用 Pandas 引擎切片并纵向写入新的 Excel
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        current_row = 0
        for i in range(0, total_cols, cols_per_block):
            # 切取当前 7 天的数据，Pandas 会自动带上最左侧的工序和设备列
            chunk_df = df.iloc[:, i: i + cols_per_block]

            # 将切片写入同一个 Sheet，startrow 控制它写在多靠下的位置
            chunk_df.to_excel(writer, sheet_name='Sheet1', startrow=current_row)

            # 计算下一次写入的起始行：当前块的数据行数 + 3行表头 + 2行空白隔离防拥挤
            current_row += len(chunk_df) + 3 + 2

    print(f"🎨 切片完成，正在对 {output_filename} 应用宽行与居中排版...")

    # 重新用 openpyxl 打开刚刚堆叠好的文件，进行最终的格式美化
    wb = openpyxl.load_workbook(output_filename)
    ws = wb.active

    # 遍历所有行和列
    for row in range(1, ws.max_row + 1):
        max_line_count = 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                # 判断：如果是前两列（设备信息），或者是表头文字（早班、周一等），则【全局居中】
                cell_text = str(cell.value)
                if col <= 2 or cell_text in ['早班', '晚班', '周一', '周二', '周三', '周四', '周五', '周六',
                                             '周日'] or '/' in cell_text:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                # 其他的排产明细数据，采用【靠左 + 垂直居中】，方便阅读长规格
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

                # 计算单元格内文字的行数，用于撑开行高
                line_count = cell_text.count('\n') + 1
                if line_count > max_line_count:
                    max_line_count = line_count

        # 动态赋予行高：文字行数 * 25 的基准系数
        ws.row_dimensions[row].height = max_line_count * 25

    # 设置列宽
    for col in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        if col <= 2:
            ws.column_dimensions[col_letter].width = 18  # 前两列设备描述稍窄
        else:
            ws.column_dimensions[col_letter].width = 30  # 排产明细数据加宽

    wb.save(output_filename)
    print(f"✅ 转换完成！请打开查看 7 天折叠版排产表: {output_filename}")


if __name__ == "__main__":
    # 输入：你用 output_dispatch.py 生成的那个完整的横向长文件
    INPUT_FILE = "210车间派工单_7月_hybrid.xlsx"

    # 输出：重组堆叠后的新文件
    OUTPUT_FILE = "210车间排产看板_7月_hybrid_7天堆叠版.xlsx"

    convert_to_stacked_format(INPUT_FILE, OUTPUT_FILE, days_per_block=7)
