import pandas as pd
import openpyxl
from openpyxl.styles import Alignment


def convert_to_stacked_format(input_filename, output_filename, days_per_block=7):
    print("=========================================")
    print("✂️ 启动看板切片与堆叠引擎")
    print("=========================================")
    print(f"📥 正在读取原始长横轴排产表: {input_filename} ...")

    try:
        # 读取原始 Excel：指定前 3 行是表头（日期、星期、班次），前 3 列是索引（工序、设备、药品规格）
        df = pd.read_excel(input_filename, header=[0, 1, 2], index_col=[0, 1, 2])
    except Exception as e:
        print(f"❌ 读取文件失败，请确保 {input_filename} 存在且格式正确。\n错误信息: {e}")
        return

    # 每天 2 个班次，计算每块截取多少列
    cols_per_block = days_per_block * 2
    total_cols = len(df.columns)

    print(f"🔄 准备按 {days_per_block} 天（{cols_per_block} 列）为一周期进行切片堆叠...")

    # 使用 Pandas 引擎切片并纵向写入新的 Excel
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        current_row = 0
        for i in range(0, total_cols, cols_per_block):
            # 1. 切取当前 7 天的数据列
            chunk_df = df.iloc[:, i: i + cols_per_block]

            # 2. 核心拦截：定义什么是“空行”（单元格为空字符串、NaN、或0的行）
            is_empty_row = chunk_df.apply(lambda row: row.astype(str).str.strip().isin(['', 'nan', '0', '0.0']).all(),
                                          axis=1)

            # 3. 只保留这 7 天内有产量的药品行，没产量的药在当前Block中直接隐藏
            chunk_df = chunk_df[~is_empty_row]

            if chunk_df.empty:
                continue

            # 4. 将精简后的切片块纵向追加写入到同一个 Worksheet 中
            chunk_df.to_excel(writer, sheet_name='Sheet1', startrow=current_row)

            # 5. 精确计算下一段分块的下落位置（动态高度：3行表头 + 实际保留的数据行数 + 2行空白隔离）
            current_row += len(chunk_df) + 3 + 2

    print(f"🎨 切片完成，正在对 {output_filename} 应用宽行与居中排版...")

    # 重新用 openpyxl 打开刚刚堆叠好的文件，进行最终的格式美化
    wb = openpyxl.load_workbook(output_filename)
    ws = wb.active

    # 遍历所有行和列
    for row in range(1, ws.max_row + 1):
        max_line_count = 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                # 判断：如果是前三列（工序、设备、药品规格），或者是表头文字（早班、周一等），则【全局居中】
                cell_text = str(cell.value)
                if col <= 3 or cell_text in ['早班', '晚班', '周一', '周二', '周三', '周四', '周五', '周六',
                                             '周日'] or '/' in cell_text:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                # 其他的排产产量纯数字数据，采用【右对齐 + 垂直居中】，美观紧凑
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='right')

                line_count = cell_text.count('\n') + 1
                if line_count > max_line_count:
                    max_line_count = line_count

        ws.row_dimensions[row].height = max_line_count * 25

    # 精准设置列宽
    for col in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        if col <= 2:
            ws.column_dimensions[col_letter].width = 16  # 工序、设备列
        elif col == 3:
            ws.column_dimensions[col_letter].width = 38  # 药品规格列加宽展示
        else:
            ws.column_dimensions[col_letter].width = 12  # 纯数字产量列

    wb.save(output_filename)
    print(f"✅ 转换完成！请打开查看 7 天折叠版排产表: {output_filename}")


if __name__ == "__main__":
    # 输入：你用 createOutput.py 生成的那个完整的横向长文件
    INPUT_FILE = "210车间全量产量派工单_7月V4_cp_1.xlsx"

    # 输出：重组堆叠后的新文件
    OUTPUT_FILE = "210车间排产看板_7月7天堆叠版_1.xlsx"

    convert_to_stacked_format(INPUT_FILE, OUTPUT_FILE, days_per_block=7)