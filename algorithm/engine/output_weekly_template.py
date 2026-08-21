import math
import re
import argparse
from collections import defaultdict
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from input_data import build_schedule_inputs, get_mix_spec, read_aps_table, split_drug_item


NAME_MAPPING = {
    "阿司匹林肠溶片（过评）": "阿司匹林肠溶片100mg",
    "左氧氟沙星片(片剂)": "左氧氟沙星片",
    "琥珀酸美托洛尔缓释胶囊(硬胶囊)": "琥珀酸美托洛尔胶囊",
    "缬沙坦氨氯地平片(片剂)": "缬沙坦氨氯地平片",
    "盐酸伊托必利片(片剂)": "盐酸伊托必利片",
    "缬沙坦氢氯噻嗪片（过评）": "缬沙坦氢氯噻嗪片",
    "瑞巴派特片(片剂)": "瑞巴派特片",
    "艾司奥美拉唑镁肠溶胶囊(硬胶囊)": "艾美",
    "盐酸二甲双胍缓释片（过评）": "盐酸二甲双胍缓释片",
    "非洛地平缓释片": "非诺地平缓释片",
}

SPEC_MAPPING = {
    "47.5mg×14粒×2板×400盒": "47.5mg×14片×2板×400盒",
    "10mg×100片×600瓶/10瓶*60包": "100片/瓶*600瓶/箱",
    "80mg×7片×4板×400盒": "7片×4板×400盒",
    "50mg×10片×2板×400盒": "10片×2板×400盒",
    "80mg×14片×2板×400盒": "（80mg:12.5mg）14片×2板×400盒",
    "0.5gx30片x400瓶": "30片/瓶×400盒/箱",
    "40mg*7粒/板*4板/盒*200盒": "40mg×7粒×4板×200盒",
    "20mg*7粒/板*4板/盒*200盒": "20mg×7粒×4板×200盒",
    "0.2g×10片×1板×400盒": "10粒×1板×400盒",
    "0.2g×10片×2板×400盒": "10粒×2板×400盒",
    "0.5mg×12粒×2板×400盒": "0.5mg×12片×2板×400盒",
    "25mg（按C21H29N6O5P计）×15片×4板×400盒": "15片×4板×400盒",
    "23.75mg×14粒×2板×400盒": "23.75mg×14片×2板×400盒",
    "5mg*10片/板*4板*400盒": "10片×4板×400盒",
}

STAGE_CONFIG = {
    1: {
        "title": "配料计划",
        "start_col": "配料开工(班时)",
        "device_col": "配料设备",
        "duration": "p1",
        "rate_col": "配料_班产量（万片）",
    },
    2: {
        "title": "压片计划",
        "start_col": "压片开工(班时)",
        "device_col": "压片设备",
        "duration": "p2",
        "rate_col": "压片_班产量",
    },
    3: {
        "title": "包衣计划",
        "start_col": "包衣开工(班时)",
        "device_col": "包衣设备",
        "duration": "p3",
        "rate_col": "包衣_班产量",
    },
    4: {
        "title": "铝塑包装计划",
        "duration": "p4",
        "rate_col": "分装/铝塑_班产量（万片）",
    },
}

HEADER = [
    "生产区域", "产品代码", "品名", "规格", "包装规格", "生产类型",
    "生产计划量", "销售计划量", "计划说明",
]


DAY_NAMES = ("周日", "周一", "周二", "周三", "周四", "周五", "周六")


def _validate_shifts_per_day(shifts_per_day):
    if isinstance(shifts_per_day, bool) or int(shifts_per_day) != shifts_per_day:
        raise ValueError("shifts_per_day must be a positive integer")
    shifts_per_day = int(shifts_per_day)
    if shifts_per_day <= 0:
        raise ValueError("shifts_per_day must be a positive integer")
    return shifts_per_day


def _build_header(shifts_per_day):
    shifts_per_day = _validate_shifts_per_day(shifts_per_day)
    return (
        HEADER[:6]
        + [
            f"{day}{shift_no}"
            for day in DAY_NAMES
            for shift_no in range(1, shifts_per_day + 1)
        ]
        + HEADER[-3:]
    )


def _safe_float(value, default=0.0):
    if value is None or pd.isna(value):
        return default
    text = str(value).strip()
    if text in {"", "-", "——", "nan", "None"}:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _clean_number(value):
    if value is None or abs(value) < 1e-9:
        return None
    value = round(float(value), 2)
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return value


def _normalize_spec(spec):
    return str(spec).replace("*", "×").replace("x", "×").strip()


def _apply_item_mapping(name, spec):
    name = str(name).strip()
    spec = str(spec).strip()
    mapped_spec = SPEC_MAPPING.get(spec, spec)
    return NAME_MAPPING.get(name, name), _normalize_spec(mapped_spec)


def _item_key(name, spec):
    mapped_name, mapped_spec = _apply_item_mapping(name, spec)
    return f"{mapped_name} {mapped_spec}"


def _read_aps_table(aps_file):
    df = read_aps_table(aps_file)
    df["包装规格"] = df["包装规格"].astype(str).str.strip().map(_normalize_spec)
    return df


def _build_stage_rates(aps_file, items):
    df = _read_aps_table(aps_file)
    rates = {stage: {} for stage in STAGE_CONFIG}
    item_set = set(items)
    for _, row in df.iterrows():
        key = f"{str(row['品种']).strip()} {str(row['包装规格']).strip()}"
        if key not in item_set:
            continue
        for stage, cfg in STAGE_CONFIG.items():
            rates[stage][key] = _safe_float(row.get(cfg["rate_col"]), 0.0)
    return rates


def _build_item_metadata(demo_file, items, department="210车间"):
    df = pd.read_excel(demo_file)
    df = df[df["部门"] == department].copy()
    df = df.dropna(subset=["存货名称", "规格"])
    df["存货名称"] = df["存货名称"].astype(str).str.strip()
    df["规格"] = df["规格"].astype(str).str.strip()
    df["APS_KEY"] = df.apply(lambda r: _item_key(r["存货名称"], r["规格"]), axis=1)

    metadata = {}
    for _, row in df.iterrows():
        key = row["APS_KEY"]
        if key not in items or key in metadata:
            continue
        metadata[key] = {
            "产品代码": row.get("物料编码", ""),
            "品名": row.get("存货名称", split_drug_item(key)[0]),
            "包装规格": _display_pack_spec(split_drug_item(key)[1]),
            "销售计划量": _clean_number(_safe_float(row.get("提报合计"), None)),
        }

    for item in items:
        if item not in metadata:
            name, spec = split_drug_item(item)
            metadata[item] = {
                "产品代码": "",
                "品名": name,
                "包装规格": _display_pack_spec(spec),
                "销售计划量": None,
            }
    return metadata


def _display_pack_spec(spec):
    spec = _normalize_spec(spec)
    spec = re.sub(r"^（[^）]*?(?:mg|g|μg|ug)[^）]*?）", "", spec, flags=re.IGNORECASE)
    spec = re.sub(
        r"^[0-9]+(?:\.[0-9]+)?\s*(?:mg|g|μg|ug)"
        r"(?:\s*[:：/]\s*[0-9]+(?:\.[0-9]+)?\s*(?:mg|g|μg|ug))?"
        r"\s*[×x*]?",
        "",
        spec,
        flags=re.IGNORECASE,
    )
    return spec.strip() or _normalize_spec(spec)


def _week_start_sunday(day):
    return day - timedelta(days=(day.weekday() + 1) % 7)


def _shift_bounds(start_date, shift_index, shifts_per_day=2):
    shifts_per_day = _validate_shifts_per_day(shifts_per_day)
    day = start_date + timedelta(days=shift_index // shifts_per_day)
    shift_no = (shift_index % shifts_per_day) + 1
    return day, shift_no


def _is_sunday_shift(start_date, shift_index, shifts_per_day=2):
    day, _shift_no = _shift_bounds(start_date, shift_index, shifts_per_day)
    return day.weekday() == 6


def _advance_to_working_time(calendar_time, start_date, shifts_per_day=2):
    current = float(calendar_time)
    while True:
        shift_idx = math.floor(current + 1e-9)
        if not _is_sunday_shift(start_date, shift_idx, shifts_per_day):
            return current
        current = shift_idx + 1


def _add_working_duration(calendar_start, working_duration, start_date, shifts_per_day=2):
    current = _advance_to_working_time(calendar_start, start_date, shifts_per_day)
    remaining = float(working_duration)
    while remaining > 1e-9:
        current = _advance_to_working_time(current, start_date, shifts_per_day)
        shift_idx = math.floor(current + 1e-9)
        shift_end = shift_idx + 1
        available = max(0.0, shift_end - current)
        step = min(remaining, available)
        current += step
        remaining -= step
        if step <= 1e-9:
            current = shift_end
    return current


def _split_task_by_week(start, end, rate, start_date, shifts_per_day=2):
    shifts_per_day = _validate_shifts_per_day(shifts_per_day)
    result = defaultdict(lambda: [0.0] * (7 * shifts_per_day))
    if end <= start or rate <= 0:
        return result

    first_shift = math.floor(start)
    last_shift = math.ceil(end) - 1
    for shift_idx in range(first_shift, last_shift + 1):
        overlap = max(0.0, min(end, shift_idx + 1) - max(start, shift_idx))
        if overlap <= 1e-9:
            continue
        day, shift_no = _shift_bounds(start_date, shift_idx, shifts_per_day)
        if day.weekday() == 6:
            continue
        week_start = _week_start_sunday(day)
        day_offset = (day - week_start).days
        if 0 <= day_offset <= 6:
            slot = day_offset * shifts_per_day + (shift_no - 1)
            result[week_start][slot] += overlap * rate
    return result


def _add_quantity(rows, stage, device, item, start, end, rate, start_date, shifts_per_day=2):
    for week_start, slots in _split_task_by_week(
        start, end, rate, start_date, shifts_per_day
    ).items():
        key = (week_start, stage, str(device), item)
        for idx, qty in enumerate(slots):
            rows[key][idx] += qty


def _collect_visual_rows(
    result_file,
    demo_file,
    aps_file,
    start_date,
    department="210车间",
    shifts_per_day=2,
):
    (
        I, J1, J2, J3, J4, B, p1, p2, p3, p4, _p5, _d, _T, _w,
        _stage_staff_limits, _clear_time_matrices, _machine_available_time,
        _release_time, _max_continuous_run,
    ) = build_schedule_inputs(
        demo_file,
        aps_file,
        department=department,
        shifts_per_day=shifts_per_day,
    )

    rates = _build_stage_rates(aps_file, I)
    metadata = _build_item_metadata(demo_file, set(I), department=department)
    durations = {1: p1, 2: p2, 3: p3, 4: p4}
    rows = defaultdict(lambda: [0.0] * (7 * shifts_per_day))

    batch_df = pd.read_excel(result_file, sheet_name="批次工序计划")
    pack_df = pd.read_excel(result_file, sheet_name="包装及摘要")

    for _, record in batch_df.iterrows():
        item = record["药品规格"]
        if item not in B:
            continue
        for stage in (1, 2, 3):
            cfg = STAGE_CONFIG[stage]
            start = _safe_float(record.get(cfg["start_col"]), None)
            device = record.get(cfg["device_col"])
            if start is None or pd.isna(device) or str(device).strip() in {"", "-", "无需包衣", "nan"}:
                continue
            duration = _safe_float(durations[stage].get((item, str(device).strip())), 0.0)
            rate = rates[stage].get(item, 0.0)
            if duration <= 0 or rate <= 0:
                continue
            end = _add_working_duration(start, duration, start_date, shifts_per_day)
            _add_quantity(
                rows,
                stage,
                str(device).strip(),
                item,
                start,
                end,
                rate,
                start_date,
                shifts_per_day,
            )

    pack_start = {
        row["药品规格"]: (row["分装铝塑设备"], _safe_float(row["包装开工(班时)"], None))
        for _, row in pack_df.iterrows()
    }
    for item in I:
        if item not in pack_start:
            continue
        device, start0 = pack_start[item]
        device = str(device).strip()
        if start0 is None or (item, device) not in p4:
            continue
        duration = _safe_float(p4[(item, device)], 0.0)
        rate = rates[4].get(item, 0.0)
        if duration <= 0 or rate <= 0:
            continue
        start = start0
        for idx, _batch in enumerate(B[item]):
            end = _add_working_duration(start, duration, start_date, shifts_per_day)
            _add_quantity(
                rows,
                4,
                device,
                item,
                start,
                end,
                rate,
                start_date,
                shifts_per_day,
            )
            start = end

    return rows, metadata


def _week_no_sunday_based(week_start):
    return int(week_start.strftime("%U")) + 1


def _first_active_slot(slots):
    for idx, qty in enumerate(slots):
        if qty > 1e-9:
            return idx
    return len(slots)


def _write_stage_section_header(ws, row, stage, department="210车间"):
    title = STAGE_CONFIG[stage]["title"]
    cell = ws.cell(row, 1, f"{department} · {title}")
    cell.font = Font(bold=True, size=14)
    ws.row_dimensions[row].height = 28
    return row + 2


def _write_block(
    ws, row, week_start, stage, block_rows, metadata,
    department="210车间", shifts_per_day=2,
):
    shifts_per_day = _validate_shifts_per_day(shifts_per_day)
    cyan = PatternFill("solid", fgColor="00FFFF")
    yellow = PatternFill("solid", fgColor="FFFF00")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    week_end = week_start + timedelta(days=6)
    ws.cell(row, 1, f"{department}周生产计划({week_start.year}年第{_week_no_sunday_based(week_start)}周)："
                    f"{week_start:%Y.%m.%d}-{week_end:%Y.%m.%d}")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    ws.cell(row, 5, STAGE_CONFIG[stage]["title"])
    ws.cell(row, 5).font = Font(bold=True, size=12)
    ws.row_dimensions[row].height = 25

    date_row = row + 1
    header_row = row + 2
    for day_idx in range(7):
        col = 7 + day_idx * shifts_per_day
        ws.cell(date_row, col, (week_start + timedelta(days=day_idx)).strftime("%Y.%m.%d"))
        ws.cell(date_row, col).fill = cyan

    for col_idx, title in enumerate(_build_header(shifts_per_day), start=1):
        cell = ws.cell(header_row, col_idx, title)
        cell.fill = cyan
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    data_start = row + 3
    current_row = data_start
    for device, item, slots in block_rows:
        info = metadata[item]
        values = [
            device,
            info.get("产品代码", ""),
            info.get("品名", split_drug_item(item)[0]),
            get_mix_spec(item),
            info.get("包装规格", _display_pack_spec(split_drug_item(item)[1])),
            "自主生产",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(current_row, col_idx, value)
            cell.alignment = Alignment(vertical="center")
            cell.border = border

        total = 0.0
        for idx, qty in enumerate(slots):
            value = _clean_number(qty)
            if value is not None:
                total += float(qty)
                cell = ws.cell(current_row, 7 + idx, value)
                cell.fill = yellow
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            else:
                ws.cell(current_row, 7 + idx).border = border

        total_col = 7 + 7 * shifts_per_day
        ws.cell(current_row, total_col, _clean_number(total))
        ws.cell(current_row, total_col + 1, info.get("销售计划量"))
        ws.cell(current_row, total_col + 2, None)
        for col_idx in (total_col, total_col + 1, total_col + 2):
            ws.cell(current_row, col_idx).border = border
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

    summary_row = current_row
    ws.cell(summary_row, 6, "计划汇总：")
    ws.cell(summary_row, 6).font = Font(bold=True)
    ws.cell(summary_row, 6).alignment = Alignment(horizontal="center", vertical="center")
    for slot_idx in range(7 * shifts_per_day):
        col = 7 + slot_idx
        total = sum(slots[slot_idx] for _device, _item, slots in block_rows)
        ws.cell(summary_row, col, _clean_number(total))
        ws.cell(summary_row, col).font = Font(bold=True)
        ws.cell(summary_row, col).alignment = Alignment(horizontal="center", vertical="center")
    total_col = 7 + 7 * shifts_per_day
    ws.cell(summary_row, total_col, _clean_number(sum(sum(slots) for _device, _item, slots in block_rows)))
    ws.cell(summary_row, total_col).font = Font(bold=True)

    for col in range(1, total_col + 3):
        ws.cell(summary_row, col).border = border
    return summary_row + 2


def generate_template_schedule_board(
    result_file,
    demo_file,
    aps_file,
    output_file,
    start_date=datetime(2026, 7, 1),
    department="210车间",
    shifts_per_day=2,
):
    shifts_per_day = _validate_shifts_per_day(shifts_per_day)
    rows, metadata = _collect_visual_rows(
        result_file,
        demo_file,
        aps_file,
        start_date,
        department=department,
        shifts_per_day=shifts_per_day,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = department
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 13, "B": 13, "C": 32, "D": 24, "E": 20, "F": 10,
    }
    for col in range(7, 7 + 7 * shifts_per_day):
        widths[get_column_letter(col)] = 12
    total_col = 7 + 7 * shifts_per_day
    widths[get_column_letter(total_col)] = 12
    widths[get_column_letter(total_col + 1)] = 12
    widths[get_column_letter(total_col + 2)] = 16
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    all_weeks = sorted({key[0] for key in rows})
    row = 1
    # 四大板块：配料 / 压片 / 包衣 / 铝塑；板块内按周、开工时间正序。
    for stage in (1, 2, 3, 4):
        stage_weeks = [
            week_start
            for week_start in all_weeks
            if any(
                wk == week_start and st == stage and any(q > 1e-9 for q in slots)
                for (wk, st, _device, _item), slots in rows.items()
            )
        ]
        if not stage_weeks:
            continue
        row = _write_stage_section_header(ws, row, stage, department=department)
        for week_start in stage_weeks:
            block_rows = []
            for (wk, st, device, item), slots in rows.items():
                if wk == week_start and st == stage and any(q > 1e-9 for q in slots):
                    block_rows.append((device, item, slots))
            block_rows.sort(
                key=lambda x: (_first_active_slot(x[2]), str(x[0]), str(x[1]))
            )
            row = _write_block(
                ws,
                row,
                week_start,
                stage,
                block_rows,
                metadata,
                department=department,
                shifts_per_day=shifts_per_day,
            )

    for rows_cells in ws.iter_rows():
        for cell in rows_cells:
            alignment = copy(cell.alignment)
            alignment.wrap_text = False
            alignment.vertical = alignment.vertical or "center"
            cell.alignment = alignment

    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    return str(output_file)


def main():
    parser = argparse.ArgumentParser(description="Generate weekly schedule visualization workbook.")
    parser.add_argument("--result", default="排产结果明细_packclear_sundayrest.xlsx")
    parser.add_argument("--demo", default="./输入/demo2.xlsx")
    parser.add_argument("--aps", default="./输入/副本APS排产信息-4.30.xlsx")
    parser.add_argument("--output", default="可排产结果可视化.xlsx")
    parser.add_argument("--start-date", default="2026-07-01")
    parser.add_argument("--department", default="210车间")
    parser.add_argument("--shifts-per-day", type=int, default=2)
    args = parser.parse_args()

    generate_template_schedule_board(
        args.result,
        args.demo,
        args.aps,
        args.output,
        datetime.strptime(args.start_date, "%Y-%m-%d"),
        department=args.department,
        shifts_per_day=args.shifts_per_day,
    )


if __name__ == "__main__":
    main()
