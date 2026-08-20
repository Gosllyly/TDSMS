from copy import copy
from io import BytesIO

from openpyxl import load_workbook

from services.excel_service import APS_TEMPLATE_FILENAME, resolve_media_template


APS_EXPORT_FIELDS = (
    "productName",
    "packageSpecification",
    "mixingLine",
    "mixingBatchQuantity",
    "mixingShiftOutput",
    "mixingWorkerCount",
    "tabletPress",
    "tabletingShiftOutput",
    "tabletingWorkerCount",
    "coatingMachine",
    "coatingShiftOutput",
    "coatingWorkerCount",
    "dividingEquipment",
    "dividingShiftOutput",
    "dividingWorkerCount",
    "packagingEquipment",
    "packagingShiftOutput",
    "manualPackagingOutput",
    "packagingWorkerCount",
    "productionCycleDays",
    "centralizedProcurement",
    "annualSales",
)


def _excel_value(field, value):
    if field == "centralizedProcurement":
        if value is None:
            return None
        return "是" if value == 1 else "否"
    if hasattr(value, "as_tuple"):
        return float(value)
    return value


def export_aps_archive(items):
    template = resolve_media_template(APS_TEMPLATE_FILENAME, ascii_prefix="APS")
    workbook = load_workbook(template)
    sheet = workbook.worksheets[0]

    style_row = [copy(sheet.cell(3, column)._style) for column in range(1, 23)]
    row_height = sheet.row_dimensions[3].height
    if sheet.max_row >= 3:
        sheet.delete_rows(3, sheet.max_row - 2)

    for row_number, item in enumerate(items, start=3):
        for column, field in enumerate(APS_EXPORT_FIELDS, start=1):
            cell = sheet.cell(row_number, column, _excel_value(field, getattr(item, field)))
            cell._style = copy(style_row[column - 1])
        sheet.row_dimensions[row_number].height = row_height

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
