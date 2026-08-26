"""将 Excel 结果转为只读 HTML 预览，尽量贴近 WPS/Excel 屏幕显示效果。"""
from __future__ import annotations

import gzip
import html
import logging
import threading
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_gen_locks: dict[str, threading.Lock] = {}
_gen_locks_guard = threading.Lock()


class ExcelPreviewError(Exception):
    """Excel 预览生成失败。"""


class _StyleRegistry:
    """将重复的内联样式折叠为 CSS class，显著缩小 HTML 体积。"""

    def __init__(self):
        self._class_by_css: dict[str, str] = {}
        self._rules: list[tuple[str, str]] = []

    def class_for(self, css: str) -> str:
        name = self._class_by_css.get(css)
        if name is None:
            name = f"c{len(self._rules)}"
            self._class_by_css[css] = name
            self._rules.append((name, css))
        return name

    def css_block(self) -> str:
        return "\n".join(f".{name}{{{css}}}" for name, css in self._rules)


def resolve_preview_path(excel_path: Path) -> Path:
    return excel_path.with_name(f"{excel_path.stem}.preview.html")


def resolve_preview_gzip_path(excel_path: Path) -> Path:
    return excel_path.with_name(f"{excel_path.stem}.preview.html.gz")


def _cache_valid(source: Path, target: Path) -> bool:
    return (
        target.is_file()
        and target.stat().st_mtime >= source.stat().st_mtime
        and target.stat().st_size > 0
    )


def _preview_cache_valid(source: Path) -> bool:
    html_path = resolve_preview_path(source)
    gzip_path = resolve_preview_gzip_path(source)
    return _cache_valid(source, html_path) and _cache_valid(source, gzip_path)


def _gen_lock(source: Path) -> threading.Lock:
    try:
        key = str(source.resolve())
    except OSError:
        key = str(source)
    with _gen_locks_guard:
        lock = _gen_locks.get(key)
        if lock is None:
            lock = threading.Lock()
            _gen_locks[key] = lock
        return lock


def excel_to_preview_html(excel_path: str | Path, *, force: bool = False) -> Path:
    source = Path(excel_path)
    if not source.is_file():
        raise ExcelPreviewError("结果文件不存在")

    target = resolve_preview_path(source)
    if not force and _preview_cache_valid(source):
        return target

    with _gen_lock(source):
        if not force and _preview_cache_valid(source):
            return target
        html_text = _render_workbook_html(source)
        target.write_text(html_text, encoding="utf-8")
        gzip_path = resolve_preview_gzip_path(source)
        gzip_path.write_bytes(gzip.compress(html_text.encode("utf-8"), compresslevel=6))
    return target


def open_preview_for_response(
    excel_path: str | Path, *, accept_gzip: bool = False
) -> tuple[Path, bool]:
    """返回预览文件路径；若可用 gzip 缓存则优先返回压缩版。"""
    source = Path(excel_path)
    preview_path = excel_to_preview_html(source)
    if accept_gzip:
        gzip_path = resolve_preview_gzip_path(source)
        if _cache_valid(source, gzip_path):
            return gzip_path, True
    return preview_path, False


_preconvert_pending: set[str] = set()
_preconvert_lock = threading.Lock()


def schedule_excel_to_preview(excel_path: str | Path) -> None:
    """结果落盘后后台预生成 HTML，缩短首次预览等待。"""
    source = Path(excel_path)
    if not source.is_file():
        return
    try:
        key = str(source.resolve())
    except OSError:
        key = str(source)
    with _preconvert_lock:
        if key in _preconvert_pending:
            return
        if _preview_cache_valid(source):
            return
        _preconvert_pending.add(key)

    def _job():
        try:
            excel_to_preview_html(source)
            logger.info("Excel HTML 预览预生成完成: %s", resolve_preview_path(source))
        except Exception:
            logger.exception("Excel HTML 预览预生成失败: %s", source)
        finally:
            with _preconvert_lock:
                _preconvert_pending.discard(key)

    threading.Thread(target=_job, name="excel-html-preconvert", daemon=True).start()


def _theme_rgb(theme: int, tint: float = 0.0) -> str | None:
    palette = {
        0: "FFFFFF",
        1: "000000",
        2: "E7E6E6",
        3: "44546A",
        4: "5B9BD5",
        5: "ED7D31",
        6: "A5A5A5",
        7: "FFC000",
        8: "4472C4",
        9: "70AD47",
    }
    base = palette.get(theme)
    if not base:
        return None
    if not tint:
        return base
    return _apply_tint(base, tint)


def _apply_tint(rgb: str, tint: float) -> str:
    rgb = rgb[-6:]
    r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)

    def _channel(c: int) -> int:
        if tint < 0:
            return max(0, min(255, int(c * (1.0 + tint))))
        return max(0, min(255, int(c + (255 - c) * tint)))

    return f"{_channel(r):02X}{_channel(g):02X}{_channel(b):02X}"


def _color_to_css(color) -> str | None:
    if color is None or getattr(color, "type", None) is None:
        return None
    ctype = color.type
    if ctype == "rgb" and color.rgb:
        rgb = str(color.rgb).upper()
        if rgb in {"00000000", "00FFFFFF", "FFFFFFFF"}:
            return None
        if len(rgb) == 8:
            rgb = rgb[2:]
        if rgb == "000000":
            return None
        return f"#{rgb[-6:]}"
    if ctype == "theme" and color.theme is not None:
        rgb = _theme_rgb(int(color.theme), float(color.tint or 0))
        return f"#{rgb}" if rgb else None
    if ctype == "indexed" and color.indexed is not None:
        idx = int(color.indexed)
        if 0 <= idx < len(COLOR_INDEX):
            raw = COLOR_INDEX[idx]
            if isinstance(raw, str) and len(raw) >= 6:
                return f"#{raw[-6:]}"
    return None


def _border_css(side) -> str | None:
    if side is None or not side.style:
        return None
    color = _color_to_css(side.color) or "#808080"
    style = side.style
    if style in {"hair", "dotted"}:
        width, css_style = "1px", "dotted"
    elif style in {"dashed", "dashDot", "dashDotDot", "mediumDashed"}:
        width, css_style = "1px", "dashed"
    elif style in {"medium", "mediumDashDot", "mediumDashDotDot"}:
        width, css_style = "2px", "solid"
    elif style in {"thick", "double"}:
        width, css_style = "3px", "solid"
    else:
        width, css_style = "1px", "solid"
    return f"{width} {css_style} {color}"


def _column_width_px(ws, col_idx: int) -> float:
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    width = dim.width if dim and dim.width is not None else 8.43
    return round(width * 7.0 + 5.0, 1)


def _row_height_px(ws, row_idx: int) -> float:
    dim = ws.row_dimensions.get(row_idx)
    if dim and dim.height is not None:
        return round(float(dim.height) * 96.0 / 72.0, 1)
    return 18.0


def _cell_value_html(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if abs(value - round(value)) < 1e-9:
            text = str(int(round(value)))
        else:
            text = f"{value:g}"
    else:
        text = str(value)
    return html.escape(text).replace("\n", "<br>")


def _build_merge_map(ws):
    """top-left -> (rowspan, colspan); covered cells -> None."""
    merge_map: dict[tuple[int, int], tuple[int, int] | None] = {}
    for merged in ws.merged_cells.ranges:
        min_row, min_col = merged.min_row, merged.min_col
        max_row, max_col = merged.max_row, merged.max_col
        merge_map[(min_row, min_col)] = (
            max_row - min_row + 1,
            max_col - min_col + 1,
        )
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    merge_map[(r, c)] = None
    return merge_map


def _cell_style_css(cell) -> str:
    styles: list[str] = ["background:#fff"]
    fill = cell.fill
    if fill and fill.fill_type == "solid":
        bg = _color_to_css(fill.fgColor) or _color_to_css(fill.start_color)
        if bg:
            styles[0] = f"background:{bg}"

    font = cell.font
    if font:
        if font.bold:
            styles.append("font-weight:700")
        if font.italic:
            styles.append("font-style:italic")
        size = font.size or 11
        styles.append(f"font-size:{size}pt")
        if font.color:
            fg = _color_to_css(font.color)
            if fg:
                styles.append(f"color:{fg}")
        family = font.name or "Calibri"
        styles.append(
            f"font-family:{html.escape(family)},'Microsoft YaHei',sans-serif"
        )
    else:
        styles.append("font-size:11pt")
        styles.append("font-family:Calibri,'Microsoft YaHei',sans-serif")

    align = cell.alignment
    if align:
        if align.horizontal in {"left", "center", "right", "justify"}:
            styles.append(f"text-align:{align.horizontal}")
        if align.vertical in {"top", "center", "bottom"}:
            styles.append(f"vertical-align:{align.vertical}")
        if align.wrap_text:
            styles.append("white-space:pre-wrap")

    border = cell.border
    if border:
        for edge, attr in (
            ("top", border.top),
            ("right", border.right),
            ("bottom", border.bottom),
            ("left", border.left),
        ):
            css = _border_css(attr)
            if css:
                styles.append(f"border-{edge}:{css}")

    return ";".join(styles)


def _render_sheet_html(ws, styles: _StyleRegistry) -> str:
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    merge_map = _build_merge_map(ws)

    col_group = [
        f'<col style="width:{_column_width_px(ws, c)}px"/>'
        for c in range(1, max_col + 1)
    ]

    rows_html: list[str] = []
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col), start=1):
        height = _row_height_px(ws, r)
        cells: list[str] = []
        for c, cell in enumerate(row, start=1):
            if (r, c) in merge_map and merge_map[(r, c)] is None:
                continue
            css_class = styles.class_for(_cell_style_css(cell))
            attrs = [f'class="{css_class}"']
            span = merge_map.get((r, c))
            if span:
                rowspan, colspan = span
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')
            cells.append(f"<td {' '.join(attrs)}>{_cell_value_html(cell.value)}</td>")
        rows_html.append(f'<tr style="height:{height}px">{"".join(cells)}</tr>')

    sheet_name = html.escape(ws.title)
    return f"""
    <div class="sheet-wrap">
      <div class="sheet-title">{sheet_name}</div>
      <table class="sheet">
        <colgroup>{''.join(col_group)}</colgroup>
        <tbody>
          {''.join(rows_html)}
        </tbody>
      </table>
    </div>
    """


def _render_workbook_html(source: Path) -> str:
    wb = load_workbook(source, data_only=True)
    try:
        styles = _StyleRegistry()
        sheets_html = [_render_sheet_html(ws, styles) for ws in wb.worksheets]
        title = html.escape(source.stem)
        return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{title}</title>
<style>
  html, body {{
    margin: 0;
    padding: 0;
    height: 100%;
    background: #cfcfcf;
    font-family: "Microsoft YaHei", "微软雅黑", Arial, sans-serif;
  }}
  .toolbar {{
    position: sticky;
    top: 0;
    z-index: 5;
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 8px 12px;
    background: #f3f3f3;
    border-bottom: 1px solid #bbb;
    font-size: 13px;
    color: #333;
  }}
  .toolbar strong {{ font-size: 14px; }}
  .stage {{
    overflow: auto;
    height: calc(100% - 38px);
    padding: 12px;
    box-sizing: border-box;
  }}
  .sheet-wrap {{
    display: inline-block;
    min-width: 100%;
    background: #fff;
    box-shadow: 0 1px 4px rgba(0,0,0,.18);
  }}
  table.sheet {{
    border-collapse: collapse;
    table-layout: fixed;
    background: #fff;
    color: #000;
  }}
  table.sheet td {{
    box-sizing: border-box;
    padding: 1px 4px;
    white-space: nowrap;
    overflow: hidden;
    vertical-align: middle;
    background: #fff;
    border: none;
  }}
  .sheet-title {{
    padding: 8px 10px;
    font-weight: 700;
    border-bottom: 1px solid #ddd;
    background: #fafafa;
  }}
  {styles.css_block()}
</style>
</head>
<body>
  <div class="toolbar">
    <strong>结果预览（只读）</strong>
    <span>{title}</span>
  </div>
  <div class="stage">
    {''.join(sheets_html)}
  </div>
</body>
</html>
"""
    finally:
        wb.close()
