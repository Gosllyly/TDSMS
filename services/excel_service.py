from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook


class ExcelValidationError(ValueError):
    pass


APS_TEMPLATE_FILENAME = "APS排程信息模板.xlsx"
PLAN_TEMPLATE_FILENAME = "药业车间分解编排计划模板.xlsx"


def resolve_media_template(preferred_name: str, *, ascii_prefix: str | None = None) -> Path:
    """解析 media/templates 下的模板文件。

    源码使用 UTF-8 中文文件名；部分 Windows 环境中磁盘文件名编码不一致，
    导致精确路径找不到。优先精确匹配，失败后按 ASCII 前缀或目录内唯一候选回退。
    """
    templates_dir = Path(settings.MEDIA_ROOT) / "templates"
    preferred = templates_dir / preferred_name
    if preferred.is_file():
        return preferred

    if not templates_dir.is_dir():
        raise FileNotFoundError(f"模板目录不存在: {templates_dir}")

    candidates = [
        path for path in templates_dir.iterdir()
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xls"}
    ]
    if not candidates:
        raise FileNotFoundError(f"模板目录为空: {templates_dir}")

    if ascii_prefix:
        matched = [path for path in candidates if path.name.startswith(ascii_prefix)]
        if len(matched) == 1:
            return matched[0]
        if matched:
            # 同前缀多个时，选体积最大的（正式模板通常更大）
            return max(matched, key=lambda path: path.stat().st_size)

    # 计划模板等无文文件名：回退为目录中非 APS 的唯一 xlsx
    if preferred_name == PLAN_TEMPLATE_FILENAME or (
        ascii_prefix is None and not preferred_name.startswith("APS")
    ):
        non_aps = [path for path in candidates if not path.name.startswith("APS")]
        if len(non_aps) == 1:
            return non_aps[0]
        if non_aps:
            return max(non_aps, key=lambda path: path.stat().st_size)

    raise FileNotFoundError(
        f"未找到模板文件: {preferred_name}（目录: {templates_dir}）"
    )


APS_HEADERS = [
    "品种", "包装规格", "配料线体", "批量（万片/粒）", "班产量（万片）", "用人",
    "压片机", "班产量", "用人", "包衣机", "班产量", "用人", "操作间及设备",
    "班产量（万片）", "用人", "操作间及设备", "班产量（万片）", "手工包装（1人产量）",
    "用人", "生产周期/天", "是否集采品种", "年销量/万",
]
PLAN_HEADERS = ["部门", "物料编码", "存货名称", "规格", "U8现存量", "月份生产计划", "提报合计"]
# 允许「月份生产计划」「7月份生产计划」「07月份生产计划」
MONTHLY_PLAN_HEADER_RE = re.compile(r"^(?:0?[1-9]|1[0-2])?月份生产计划$")


def _is_plan_header_row(headers):
    if len(headers) < 7:
        return False
    expected = PLAN_HEADERS[:]
    actual = headers[:7]
    if actual[5] and MONTHLY_PLAN_HEADER_RE.match(actual[5]):
        actual = actual[:]
        actual[5] = expected[5]
    return actual == expected


def _load_rows(uploaded_file):
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        raise ExcelValidationError("仅支持.xlsx或.xls格式文件")
    content = uploaded_file.read()
    uploaded_file.seek(0)
    if suffix == ".xls":
        try:
            import xlrd
            book = xlrd.open_workbook(file_contents=content)
            sheet = book.sheet_by_index(0)
            return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        except Exception as exc:
            raise ExcelValidationError("无法读取.xls文件，请检查文件是否损坏") from exc
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    except Exception as exc:
        raise ExcelValidationError("无法读取Excel文件，请检查文件是否损坏") from exc


def _text(value):
    if value is None:
        return None
    value = str(value).strip()
    return None if value in {"", "-", "—", "——"} else value


def _decimal(value, row, column):
    value = _text(value)
    if value is None:
        return None
    if value in {"#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"}:
        return None
    try:
        result = Decimal(value)
    except (InvalidOperation, ValueError) as exc:
        raise ExcelValidationError(f"第{row}行：{column}必须为数字") from exc
    if result < 0:
        raise ExcelValidationError(f"第{row}行：{column}不能为负数")
    return result


def _integer(value, row, column):
    result = _decimal(value, row, column)
    if result is None:
        return None
    if result != result.to_integral_value():
        raise ExcelValidationError(f"第{row}行：{column}必须为整数")
    return int(result)


def parse_aps_file(uploaded_file):
    rows = _load_rows(uploaded_file)
    if len(rows) < 3:
        raise ExcelValidationError("APS文件没有可导入的数据")
    first = [_text(v) for v in rows[0][:22]]
    second = [_text(v) for v in rows[1][:22]]
    combined = [second[index] or first[index] for index in range(22)]
    if combined != APS_HEADERS:
        raise ExcelValidationError("APS文件表头与系统模板不一致")
    result = []
    for row_no, row in enumerate(rows[2:], start=3):
        values = list(row[:22]) + [None] * max(0, 22 - len(row))
        if not any(_text(v) for v in values):
            continue
        product, specification = _text(values[0]), _text(values[1])
        if not product or not specification:
            raise ExcelValidationError(f"第{row_no}行：品种和包装规格不能为空")
        centralized = _text(values[20])
        if centralized not in {"是", "否", "1", "0"}:
            raise ExcelValidationError(f"第{row_no}行：是否集采品种只能填写是或否")
        result.append({
            "productName": product, "packageSpecification": specification,
            "mixingLine": _text(values[2]), "mixingBatchQuantity": _decimal(values[3], row_no, "配料批量"),
            "mixingShiftOutput": _decimal(values[4], row_no, "配料班产量"), "mixingWorkerCount": _integer(values[5], row_no, "配料用人"),
            "tabletPress": _text(values[6]), "tabletingShiftOutput": _decimal(values[7], row_no, "压片班产量"),
            "tabletingWorkerCount": _integer(values[8], row_no, "压片用人"), "coatingMachine": _text(values[9]),
            "coatingShiftOutput": _decimal(values[10], row_no, "包衣班产量"), "coatingWorkerCount": _integer(values[11], row_no, "包衣用人"),
            "dividingEquipment": _text(values[12]), "dividingShiftOutput": _decimal(values[13], row_no, "分装班产量"),
            "dividingWorkerCount": _integer(values[14], row_no, "分装用人"), "packagingEquipment": _text(values[15]),
            "packagingShiftOutput": _decimal(values[16], row_no, "包装班产量"), "manualPackagingOutput": _decimal(values[17], row_no, "手工包装产量"),
            "packagingWorkerCount": _integer(values[18], row_no, "包装用人"), "productionCycleDays": _decimal(values[19], row_no, "生产周期"),
            "centralizedProcurement": 1 if centralized in {"是", "1"} else 0,
            "annualSales": _decimal(values[21], row_no, "年销量"),
        })
    if not result:
        raise ExcelValidationError("APS文件没有可导入的数据")
    return result


def parse_plan_file(uploaded_file):
    rows = _load_rows(uploaded_file)
    if not rows or not _is_plan_header_row([_text(v) for v in rows[0][:7]]):
        raise ExcelValidationError("计划文件表头与系统模板不一致")
    result = []
    for row_no, row in enumerate(rows[1:], start=2):
        values = list(row[:7]) + [None] * max(0, 7 - len(row))
        if not any(_text(v) for v in values):
            continue
        core = [_text(v) for v in values[:4]]
        # 模板末尾可能存在格式残留空行；没有物料编码的行不属于业务数据。
        if not core[1]:
            continue
        if not core[0] or not core[2]:
            raise ExcelValidationError(f"第{row_no}行：部门和存货名称不能为空")
        result.append({
            "departmentName": core[0], "materialCode": core[1], "inventoryName": core[2], "specification": core[3] or "",
            "u8CurrentStock": _decimal(values[4], row_no, "U8现存量"),
            "monthlyProductionPlan": _decimal(values[5], row_no, "月份生产计划"),
            "submittedTotal": _decimal(values[6], row_no, "提报合计"),
        })
    if not result:
        raise ExcelValidationError("计划文件没有可导入的数据")
    return result
